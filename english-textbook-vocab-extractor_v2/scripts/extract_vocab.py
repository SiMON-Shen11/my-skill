# -*- coding: utf-8 -*-
"""
英语教材 PDF 词汇表提取脚本（通用版）。

用法:
  python extract_vocab.py <PDF路径> [选项]

选项:
  --pages START-END     词汇表所在页码范围（1-based, 闭区间），默认自动扫描全 PDF
  --cut X               双栏切分 x 坐标，默认 auto（按页动态检测）
  --top-skip Y          页眉跳过高度，默认 52.0
  --output PATH         输出 Excel 路径，默认 ./词汇表.xlsx
  --unit-header TEXT    单元标题标记关键词，默认 "Unit"
  --json PATH           同时输出中间 JSON，默认不输出
  --scan-only           仅扫描定位词汇表页码，不提取

示例:
  python extract_vocab.py "/path/to/textbook.pdf" --pages 106-114 --output vocab.xlsx
  python extract_vocab.py "/path/to/textbook.pdf" --scan-only
"""
import sys, os, re, json, argparse
from collections import Counter

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed. Run: pip install pdfplumber openpyxl", file=sys.stderr)
    sys.exit(1)

# ==================== 常量 ====================
Y_BUCKET = 4.0       # 行聚类容差
LINE_TOL = 6.0       # 同行容差
HW_PAD = 30.0       # 词头 x0 窗口宽度

JUNK = {"PB", "Page", "Words", "Expressions", "Each", "Unit",
        "关注", "微信", "公众号", "捷思课堂", "获取更多学习资料"}

POS_ALLOWED = {"n.", "v.", "adj.", "adv."}
POS_CANON = {
    "n.": "n.", "v.": "v.", "adj.": "adj.", "adv.": "adv.",
    "n": "n.", "v": "v.", "adj": "adj.", "adv": "adv.",
    "modal v.": "modal v.", "link v.": "link v.",
    "modal v": "modal v.", "link v": "link v.",
    "prep.": "prep.", "pron.": "pron.", "det.": "det.", "conj.": "conj.",
    "interj.": "interj.", "num.": "num.", "art.": "art.", "abbr.": "abbr.",
    "aux.": "aux.", "sing.": "sing.", "pl.": "pl.", "int.": "int.",
    "phr.": "phr.", "sym.": "sym.", "etc.": "etc.",
}

POS_RE = re.compile(
    r'(?:modal|link)\s+v\.?'
    r'|n\.|v\.|adj\.|adv\.|prep\.|pron\.|det\.|conj\.|interj\.|num\.|art\.|abbr\.|aux\.'
    r'|sing\.|pl\.|int\.|phr\.|sym\.|etc\.'
)


# ==================== 工具函数 ====================
def is_cjk_char(c):
    o = ord(c)
    if o <= 127:
        return False
    if 0xE000 <= o <= 0xF8FF:
        return False
    if 0x2018 <= o <= 0x201F:
        return False
    if o == 0x2026:  # … 水平省略号（英文短语占位符，如 connect … with），非中文释义
        return False
    return True


def has_pua(tok):
    return any(0xE000 <= ord(c) <= 0xF8FF for c in tok)


def decode_pua(tok):
    return ''.join(chr(ord(c) - 0xF000) if 0xF000 <= ord(c) <= 0xF0FF else c
                   for c in tok)


# 音标特征字符：ASCII 编码音标常用特殊符号/大写字母表示音素
# : 长音、@ 央元音、{} 元音、I ɪ、U ʊ、Q ɒ、V ʌ、N ŋ、S ʃ、T θ、D ð、Z ʒ、3 ɜ
PHONETIC_MARKERS = re.compile(r'[:@{}]|[IUVQNSDTZ3]')


def is_ascii_phonetic(t):
    """检测 ASCII 音标 token（如 /sti:l/ 或 /st@Ul/ 或 /b3:(r)n/）。
    部分教材（如九年级全一册）使用自定义 ASCII 字体编码音标，
    而非 PUA 私有区编码。支持跨行拆分的音标片段
    （如 'kA:nv@rseISn/'、'taIm/'、'/pA:(r)t'）。"""
    if not t:
        return False
    if has_pua(t):
        return True  # PUA 音标也归为音标
    if any(is_cjk_char(c) for c in t):
        return False  # 含中文的混合 token 由 parse_entry 单独处理
    stripped = t.strip()
    if '/' in stripped:
        # 含音标特征字符（ASCII 音标编码特有符号）
        if PHONETIC_MARKERS.search(stripped):
            return True
        # 以 / 开头或结尾的不完整音标片段
        if stripped.startswith('/') or stripped.endswith('/'):
            return True
        # 成对音标（纯音标字符，如 /sti:l/ 已在上面命中，此为兜底）
        if stripped.startswith('/') and stripped.endswith('/'):
            return True
    return False


def canonical_pos(tok):
    if tok.strip().lower().startswith('modal'):
        return "modal v."
    if tok.strip().lower().startswith('link'):
        return "link v."
    return POS_CANON.get(tok.strip().lower(), tok.strip())


def is_pos_tok(text):
    if text.lower() in ("modal", "link"):
        return True
    return POS_RE.fullmatch(text) is not None


def is_headword_start(text, x0, is_ll, col_min_x, hw_thresh):
    if not is_ll:
        return False
    if not (col_min_x - 2.0 <= x0 <= hw_thresh):
        return False
    if re.search(r'[一-鿿]', text):
        return False
    if text.lower() in JUNK:
        return False
    if re.fullmatch(r'p\.\d+', text, re.I):
        return False
    if is_pos_tok(text):
        return False
    # 排除词性标记的全角括号粘连形式（如 n（.、v（.）
    _strip_cn = re.sub(r'[（），。；：！？、…—’‘]', '', text)
    if is_pos_tok(_strip_cn) or re.match(r'^(?:n|v|adj|adv|prep|pron|conj|art|num|aux)\.', _strip_cn):
        return False
    if text == '&':
        return False
    if not re.match(r"[A-Za-z]", text):
        return False
    # 排除以右括号结尾的跨行续词（如 "burned)"、"overcome)"）
    # 这些是词形变化括号的闭合部分，不是新词头
    if text.endswith(')') and '(' not in text:
        return False
    # 排除 ASCII/PUA 音标 token，但允许"英文词+音标"粘连形式作词头
    # （如 dining/daInIN/ 是 dining hall 的词头+音标粘连，应作词头）
    if is_ascii_phonetic(text) and not (has_pua(text) and re.match(r'[A-Za-z]', text)):
        return False
    # 排除含音标斜杠的 token（如 "steal" 后面紧跟 "/sti:l/"），
    # 但允许"英文词+音标"粘连形式（如 dining/daInIN/）作词头
    if '/' in text and not any(is_cjk_char(c) for c in text) \
       and not (has_pua(text) and re.match(r'[A-Za-z]', text)):
        return False
    return True


def parse_entry(tokens):
    cleaned = [t for t in tokens if not re.fullmatch(r'p\.\d+', t, re.I)]

    # 预处理：将非音标 token 中的半角括号拆分为独立 token
    # 如 '(stole' → '(' + 'stole'，'stolen)' → 'stolen' + ')'
    # 含 / 的 token（音标/音标+中文混合）整体保留，避免拆分音标内 (r)，
    # 也不误删 (sb) 等正常占位符（(sb) 在音标上下文外，括号分支正常处理）
    expanded = []
    for t in cleaned:
        if '/' in t:
            expanded.append(t)
            continue
        # 全角左括号后接英文 → 归一为半角（如 '（pl.' → '(pl.'、
        # '（=centre' → '(=centre'），使括号分支能正确处理词形标记；
        # 全角左括号后接中文（如 '（表示意愿）'）保持全角（释义括号）
        if t.startswith('（') and len(t) > 1 and not is_cjk_char(t[1]):
            t = '(' + t[1:]
        parts = re.split(r'([()])', t)
        for p in parts:
            if p:
                expanded.append(p)
    cleaned = expanded

    english_parts = []
    english_closed = False
    senses = []
    cur_pos = None
    cur_chars = []

    def flush_sense():
        if cur_pos is not None or cur_chars:
            senses.append((cur_pos, ''.join(cur_chars)))

    paren_depth = 0
    pending_paren = False
    for _idx, t in enumerate(cleaned):
        # 丢弃音标可选音片段（如 (r)、(rn)、(m)）——单/双小写音标字符括号
        # 这是音标内 (r) 因跨行/切分残留的 token，不是词形变化括号，
        # 也不是 (sb) 等占位符（占位符走括号分支正常处理）
        if re.fullmatch(r'\([rnm]{1,2}\)', t):
            continue
        if t.lower() in ("modal", "link"):
            flush_sense()
            cur_pos = "modal v." if t.lower() == "modal" else "link v."
            cur_chars = []
            continue
        if t.strip().lower() in ("v.", "v") and cur_pos in ("modal v.", "link v."):
            continue
        if t.strip() in ('(', ')', '[', ']'):
            if t.strip() == '(':
                paren_depth += 1
            else:
                paren_depth = max(0, paren_depth - 1)
            english_parts.append(t.strip())
            continue
        _stripped = re.sub(r'[（），。；：！？、…—’‘]', '', t)
        if POS_RE.fullmatch(t) or POS_RE.fullmatch(_stripped):
            pos_tok = canonical_pos(t if POS_RE.fullmatch(t) else _stripped)
            # 括号内的缩写标记（如 (pl.)、(sing.)）保留为英文，不作词性
            if paren_depth > 0 and t in ('pl.', 'sing.', 'abbr.', 'etc.', 'int.', 'pl', 'sing'):
                english_parts.append(t)
                continue
            # 词性+全角左括号粘连（如 'v（.'、'v.（'）：
            # 看下一 token——英文/音标 → 词形变化开括号（如 deal v.（dealt)）；
            # 中文 → 释义括号（如 Scrooge n.（非正式）吝啬鬼），"（"加入释义
            meaning_paren = False
            if '（' in t:
                nxt = cleaned[_idx + 1] if _idx + 1 < len(cleaned) else ''
                if not re.search(r'[一-鿿]', nxt):
                    english_parts.append('(')
                    paren_depth += 1
                else:
                    meaning_paren = True
            flush_sense()
            cur_pos = pos_tok
            cur_chars = []
            if meaning_paren:
                cur_chars.append('（')
            continue
        if has_pua(t):
            dt = decode_pua(t)
            if not english_closed:
                if not dt.startswith('('):
                    m = re.match(r"([A-Za-z][A-Za-z\-’']*)\s*/", dt)
                    if m:
                        english_parts.append(m.group(1))
                pm = re.search(r'\(([A-Za-z][A-Za-z\-’]*)(?=[/\)]|$)', dt)
                if pm and (pm.start() == 0 or dt[pm.start() - 1] == '/'):
                    english_parts.append('(' + pm.group(1))
                    paren_depth += 1
                    if ')' in dt[pm.end():]:
                        english_parts.append(')')
                        paren_depth = max(0, paren_depth - 1)
                    else:
                        pending_paren = True
                elif re.search(r'\)\s*[一-鿿]?$', dt):
                    english_parts.append(')')
                    paren_depth = max(0, paren_depth - 1)
                    pending_paren = False
                # 处理音标+词性粘连（如 /hir/v.）：移除 /.../ 后检查剩余是否为词性
                pua_phon_m = re.match(r'^/[^/]*/', dt)
                if pua_phon_m:
                    rest = dt[pua_phon_m.end():].strip()
                    if rest:
                        _stripped = re.sub(r'[（），。；：！？、…—’‘]', '', rest)
                        if POS_RE.fullmatch(rest) or POS_RE.fullmatch(_stripped):
                            pos_tok = canonical_pos(rest if POS_RE.fullmatch(rest) else _stripped)
                            flush_sense()
                            cur_pos = pos_tok
                            cur_chars = []
            cjk = ''.join(c for c in t if is_cjk_char(c))
            if cjk:
                cur_chars.append(cjk)
                english_closed = True
            continue
        # ASCII 音标 + 释义混合 token 处理
        # 部分教材（如九年级全一册）不使用 PUA 编码，
        # 而是用自定义 ASCII 字体编码音标
        # 典型：/st@Ul@n/)偷；窃取  或  /b3:(r)nd/,
        if is_ascii_phonetic(t) or ('/' in t and any(is_cjk_char(c) for c in t)):
            if not english_closed:
                # 从音标中提取括号内的词形变化（如 (stole, stolen)）
                # 注意排除 (r) 这种音标内的可选音（单/双小写字母）
                pm = re.search(r'\(([A-Za-z][A-Za-z\-\']{2,})', t)
                if pm and pm.group(1) != 'r)':  # 排除 (r) 音标
                    english_parts.append('(' + pm.group(1))
                    paren_depth += 1
                    if ')' in t[pm.end():]:
                        english_parts.append(')')
                        paren_depth = max(0, paren_depth - 1)
                    else:
                        pending_paren = True
                elif paren_depth > 0:
                    # 音标末尾的词形括号闭合（如 /st@Ul@n/)偷；窃取、
                    # /li:vz/）、/rVN/））。音标内部 (r) 的右括号不闭合
                    # （不以 )/） 结尾且后无中文时不触发）
                    if re.search(r'[）)]\s*$', t) or re.search(r'[）)][一-鿿]', t):
                        english_parts.append(')')
                        paren_depth = max(0, paren_depth - 1)
                # 处理音标+词性粘连（如 /helpfl/adj.、/krIsm@s/n.）
                # 移除第一个 /.../ 音标部分，检查剩余是否为词性
                ascii_phon_m = re.match(r'^/[^/]*/', t)
                if ascii_phon_m:
                    rest = t[ascii_phon_m.end():].strip()
                    if rest:
                        _stripped = re.sub(r'[（），。；：！？、…—’‘]', '', rest)
                        if POS_RE.fullmatch(rest) or POS_RE.fullmatch(_stripped):
                            pos_tok = canonical_pos(rest if POS_RE.fullmatch(rest) else _stripped)
                            flush_sense()
                            cur_pos = pos_tok
                            cur_chars = []
            # 提取混合 token 中的中文释义
            # ASCII音标token末尾的全角右括号）是词形变化括号闭合（如 /li:vz/）叶），非释义内容；
            # 全角左括号（可能是释义括号开始，保留
            cjk = ''.join(c for c in t if is_cjk_char(c) and c != '）')
            if cjk:
                cur_chars.append(cjk)
                english_closed = True
            continue
        cjk = ''.join(c for c in t if is_cjk_char(c))
        if cjk:
            m = POS_RE.match(t)
            if m and m.end() <= 4:
                prefix = m.group(0)
                if cur_pos in ("modal v.", "link v.") and prefix in ("v.", "v"):
                    cur_chars.append(cjk)
                    continue
                flush_sense()
                cur_pos = canonical_pos(prefix)
                cur_chars = [cjk]
                continue
            # 英文+全角左括号（如 'rather（'、'look up（'、'of（'）：英文续词，
            # （ 是释义括号（如 （非正式）、（在词典...），加入释义，不拆英文括号
            if '（' in t and re.match(r'^[A-Za-z]', t) and not re.search(r'[一-鿿]', t):
                m_open = re.match(r'^([A-Za-z][A-Za-z\-\.]*)（', t)
                if m_open:
                    if not english_parts or english_parts[-1] != m_open.group(1):
                        english_parts.append(m_open.group(1))
                    # 全角左括号是释义括号，加入释义（修复：原代码丢弃了"（"）
                    cur_chars.append('（')
                    rest = t[m_open.end():]
                    cjk_rest = ''.join(c for c in rest if is_cjk_char(c))
                    if cjk_rest:
                        cur_chars.append(cjk_rest)
                    english_closed = True
                    continue
            # 英文+全角右括号混合（如 'USA）美国；'、'rather）'）：
            # paren_depth>0 时是词形括号闭合（) 进英文）；
            # paren_depth==0 时是释义右括号（) 进释义）
            if '）' in t and re.match(r'^[A-Za-z]', t):
                m_mix = re.match(r'^([A-Za-z][A-Za-z\-\.]*)[）]?(.*)$', t)
                eng_part = m_mix.group(1)
                if paren_depth > 0:
                    # 词形闭合：英文续词（去重）+ ) 闭合
                    if not english_parts or english_parts[-1] != eng_part:
                        english_parts.append(eng_part)
                    english_parts.append(')')
                    paren_depth = max(0, paren_depth - 1)
                    # 修复：词形闭合时 rest 中的全角右括号是词形括号闭合，不应进入释义；
                    # 后续中文是新义项（如 glasses眼镜）→ 已有释义时加分号分隔
                    rest = m_mix.group(2)
                    rest_cjk = ''.join(c for c in rest if is_cjk_char(c) and c != '）')
                    if rest_cjk:
                        if cur_chars:
                            cur_chars.append('；')
                        cur_chars.append(rest_cjk)
                else:
                    # 释义右括号：英文续词（去重），) 并入释义
                    if not english_parts or english_parts[-1] != eng_part:
                        english_parts.append(eng_part)
                    full_cjk = ''.join(c for c in t if is_cjk_char(c))
                    if full_cjk:
                        cur_chars.append(full_cjk)
                english_closed = True
                continue
            cur_chars.append(cjk)
            english_closed = True
            continue
        if english_closed and paren_depth == 0:
            continue
        norm = t.replace('\u2019', "'").replace('\u2018', "'")
        if norm.strip() in ("...", "…"):
            english_parts.append("...")
            continue
        # 全角右括号归一为半角（闭合词形括号，如 "the USA）" → "the USA)"）
        # 避免全角 ） 未被识别导致括号不平衡/跨行大合并
        if '）' in norm or '（' in norm:
            norm = norm.replace('）', ')').replace('（', '(')
        # '=' 是括号内"亦作"标记（如 (= centre)），应保留为英文部分
        if re.search(r'[A-Za-z]', norm) or norm.strip() == '=':
            english_parts.append(norm)
            continue
        english_closed = True
    flush_sense()

    english = re.sub(r"\s+", " ", " ".join(english_parts)).strip()
    english = re.sub(r'\s*\(\s*', ' (', english)
    english = re.sub(r'\s*\)\s*', ') ', english).strip()
    english = re.sub(r'\s+', ' ', english).strip()

    # 人名缩写重复去重：释义行会重复词头行的姓名缩写
    # 典型：'J. K. Rowling J. K.' → 'J. K. Rowling'
    # 释义行排印为 "J. K. 罗琳（英国作家）"，J. K. 与词头重复
    m = re.search(r'\s+((?:[A-Z]\.\s*)+)$', english)
    if m:
        dup = m.group(1).strip()
        prefix = english[:m.start()].strip()
        if dup in prefix:
            english = prefix

    # 清理空括号：半角括号内是中文注释时（如 (表示方式)、泰语(的)；），
    # 括号被拆进英文产生空括号 '()'，删除之（如 'by ()' → 'by'）
    english = re.sub(r'\(\s*\)', '', english).strip()
    english = re.sub(r'\s+', ' ', english).strip()

    meaning = "；".join(chars for _, chars in senses if chars)

    # 通用后处理：英文以不完整短语结尾且释义以全角右括号开头 → 补全英文闭合
    # 典型：跨行超长条目 the United States (of America) (abbr. the US, the USA）
    #   释义开头残留 ），需闭合英文并清除释义开头的全角右括号
    if english.endswith(', the') and meaning.startswith('）'):
        english = english + ')'
        meaning = meaning[1:]

    # 通用后处理：释义含未闭合的全角左括号 → 补全右括号
    # 典型：children 释义「儿童（单数形式为」→ 补「）」
    if meaning and meaning.count('（') > meaning.count('）'):
        meaning = meaning + '）'

    all_pos = [p for p, _ in senses if p]
    if not all_pos:
        pos_out = "其它"
    else:
        ordered_unique = []
        for p in all_pos:
            if p not in ordered_unique:
                ordered_unique.append(p)
        if all(p in POS_ALLOWED for p in ordered_unique):
            pos_out = " & ".join(ordered_unique)
        else:
            pos_out = "其它"
    return english, pos_out, meaning


# ==================== 扫描定位 ====================
# 词汇表页识别的三个判断条件（教材附录词汇表通用特征）：
#   1. 位置条件：词汇表一般位于教材**倒数前 50 页**内（附录/词表区）
#   2. 结构条件：词汇表内容按单元（Unit）分组，页面含 "Unit N" 标题行
#   3. 标题条件：章节通常以 "Words and Expressions in Each Unit" 等标题开篇
VOCAB_LAST_N_PAGES = 50  # 词汇表一般在教材末尾 50 页内

# 标题模式（按优先级分两组：优先标题 = 明确的按单元词汇表章节）
VOCAB_PRIMARY_TITLE_PATTERNS = [
    r'Words\s+and\s+Expressions\s+in\s+Each\s+Unit',
    r'Vocabulary\s+in\s+Each\s+Unit',
]
VOCAB_GENERIC_TITLE_PATTERNS = [
    r'Words\s+and\s+Expressions',
    r'Word\s+List',
    r'Vocabulary\s+List',
    r'Vocabulary',  # 兜底，可能命中索引表（由 EXCLUDE 拦截）
    r'词汇表',
    r'单词表',
    r'生词表',
]

# 需要排除的页面模式（索引表等非按单元词汇表）
VOCAB_EXCLUDE_PATTERNS = [
    r'Vocabulary\s+Index',
    r'Index',
]


def _score_vocab_page(text, pno, total_pages, unit_header="Unit"):
    """按三个判断条件给单页打分，返回 (score, has_title, has_unit)。

    评分（三条件加权）：
    - 标题条件（必需）：优先标题 +6 / 通用标题 +3；
      **无标题页不得独立入选**（正文页眉 "UNIT N" 会干扰），
      只能作为标题段两侧的紧邻扩展（见 scan_for_vocab_pages）
    - 位置条件：位于教材倒数前 50 页内 +2
    - 结构条件：含 "Unit N" 分组标题 +1
    - 排除条件：命中 Index/Vocabulary Index 索引表 → -1000（直接剔除）
    """
    if any(re.search(p, text, re.IGNORECASE) for p in VOCAB_EXCLUDE_PATTERNS):
        return -1000, False, False
    has_title = False
    score = 0
    if any(re.search(p, text, re.IGNORECASE) for p in VOCAB_PRIMARY_TITLE_PATTERNS):
        score += 6
        has_title = True
    elif any(re.search(p, text, re.IGNORECASE) for p in VOCAB_GENERIC_TITLE_PATTERNS):
        score += 3
        has_title = True
    # 结构：含 "Unit N" 分组（对标题页加分，对无标题页是扩展资格）
    unit_re = re.compile(r'\b' + re.escape(unit_header) + r'\s+\d+', re.IGNORECASE)
    has_unit = bool(unit_re.search(text))
    if has_title:
        # 位置：倒数前 50 页
        if pno >= total_pages - VOCAB_LAST_N_PAGES + 1:
            score += 2
        if has_unit:
            score += 1
    return score, has_title, has_unit


def scan_for_vocab_pages(pdf_path, unit_header="Unit"):
    """扫描 PDF 定位词汇表所在页码范围。返回 (start, end) 1-based 闭区间。

    结合三个判断条件综合评分：
      1. 位置：词汇表一般在教材**倒数前 50 页**内；
      2. 结构：词汇表内容按 Unit 分组（页面含 "Unit N" 标题行）；
      3. 标题：通常以 "Words and Expressions in Each Unit" 等标题开篇。

    算法：
    1. 标题页（含优先/通用标题，非索引）作为候选主干，聚成连续段；
    2. 每段向两侧**紧邻扩展**：吸收含 "Unit N" 结构的无标题页
       （限制扩展宽度，避免正文页眉 UNIT N 无限延伸）；
    3. 段得分 = 标题页得分和 + 扩展页×0.5，取最高分段。
    """
    sigs = []
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            s, ht, hu = _score_vocab_page(text, i + 1, total, unit_header)
            sigs.append((i + 1, s, ht, hu))

    # 1. 标题页连续段（主干）
    title_segs = []
    cur = []
    for pno, s, ht, hu in sigs:
        if ht and s > 0:
            cur.append(pno)
        else:
            if cur:
                title_segs.append(cur)
                cur = []
    if cur:
        title_segs.append(cur)

    best = None
    best_score = -1
    MAX_EXTEND = 8  # 每侧最多扩展页数（正文页眉 UNIT N 防止无限延伸）
    for seg in title_segs:
        lo, hi = seg[0], seg[-1]
        seg_score = sum(sigs[p - 1][1] for p in seg)
        # 2. 向两侧扩展：紧邻且含 Unit N 的无标题页
        ext_lo = lo
        for p in range(lo - 1, lo - 1 - MAX_EXTEND, -1):
            if p < 1:
                break
            _pno, s, ht, hu = sigs[p - 1]
            if ht or not hu or s < 0:
                break
            ext_lo = p
            seg_score += 0.5
        ext_hi = hi
        for p in range(hi + 1, hi + 1 + MAX_EXTEND):
            if p > total:
                break
            _pno, s, ht, hu = sigs[p - 1]
            if ht or not hu or s < 0:
                break
            ext_hi = p
            seg_score += 0.5
        if seg_score > best_score:
            best_score = seg_score
            best = (ext_lo, ext_hi)

    if best is None:
        return None, None
    return best


def detect_column_cut(pdf_path, page_idx):
    """自动检测双栏切分 x 坐标（取页面宽度中部附近）。"""
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_idx]
        return page.width / 2.0


def detect_page_cut(words, top_skip=52.0):
    """按页动态检测双栏切分 x 坐标。

    不同页码的双栏位置可能不同（如奇数页/偶数页右栏起始 x0 不同），
    固定 cut 会导致右栏词头被误分到左栏。

    算法：
    1. 收集词头候选 token（英文开头、非词性、非JUNK、非音标、非页码）的 x0；
    2. 词头 x0 呈左右两簇，两簇间最大间隙即左右栏分界；
    3. 右簇最小 x0 = R_head_min，cut = (左栏所有token最大x0 + R_head_min) / 2。
    """
    body = [w for w in words if w['top'] >= top_skip]
    if not body:
        return 269.0

    hw_xs = []
    for w in body:
        t = w['text'].strip()
        if not re.match(r'[A-Za-z]', t):
            continue
        if re.fullmatch(r'p\.\d+', t, re.I):
            continue
        if t.startswith('/') or t.endswith(')'):
            continue
        if t in JUNK:
            continue
        # 排除词性标记（含粘连形式如 n.病人、adj.有耐心的）
        if is_pos_tok(t) or re.match(r'^(?:n|v|adj|adv|prep|pron|conj|art|num|aux)\.', t):
            continue
        hw_xs.append(w['x0'])

    hw_xs.sort()
    if len(hw_xs) < 6:
        return max(w['x1'] for w in body) / 2.0

    # 左右两簇间的最大间隙
    max_gap = 0
    split = None
    for a, b in zip(hw_xs, hw_xs[1:]):
        if b - a > max_gap:
            max_gap = b - a
            split = (a + b) / 2.0
    if split is None:
        return max(w['x1'] for w in body) / 2.0

    right_head_min = min(x for x in hw_xs if x > split)
    left_all_max = max((w['x0'] for w in body if w['x0'] < right_head_min),
                       default=right_head_min - 1.0)
    return (left_all_max + right_head_min) / 2.0


# ==================== 主提取逻辑 ====================
def extract(pdf_path, page_start, page_end, cut, top_skip, unit_header="Unit"):
    """提取词汇表，返回 entries 列表。

    cut 为 None 时按页动态检测双栏切分线。
    Unit 标题通过整页预扫描定位（按 top 归属），不依赖栏内相邻。
    """
    vocab_pages = range(page_start - 1, page_end)  # 转 0-indexed
    unit_re = re.compile(r'^' + re.escape(unit_header) + r'$', re.IGNORECASE)

    ordered = []
    unit_boundaries = {}  # pno -> [(top, unit_num)] 按 top 升序

    for pno in vocab_pages:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[pno]
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)

        # 1. 预扫描本页 Unit 标题（整页范围，跨栏定位）
        bounds = []
        for i, w in enumerate(words):
            if unit_re.match(w['text'].strip()):
                for j in range(i + 1, min(i + 4, len(words))):
                    if (re.fullmatch(r'\d+', words[j]['text'].strip())
                            and abs(words[j]['top'] - w['top']) <= LINE_TOL):
                        bounds.append((w['top'], int(words[j]['text'].strip())))
                        break
        bounds.sort()
        unit_boundaries[pno] = bounds

        # 2. 本页切分线（未指定时按页动态检测）
        page_cut = detect_page_cut(words, top_skip) if cut is None else cut

        # 3. 双栏切分
        for col in ([w for w in words if w['x0'] < page_cut],
                    [w for w in words if w['x0'] >= page_cut]):
            if not col:
                continue
            body = [w for w in col if w['top'] >= top_skip]
            if not body:
                continue
            col_min_x = min(w['x0'] for w in body)
            hw_thresh = col_min_x + HW_PAD
            toks_sorted = sorted(body, key=lambda w: w['top'])
            lines = []
            cur = [toks_sorted[0]]
            for w in toks_sorted[1:]:
                if w['top'] - cur[-1]['top'] > LINE_TOL:
                    lines.append(cur)
                    cur = [w]
                else:
                    cur.append(w)
            lines.append(cur)
            for line in lines:
                grp = sorted(line, key=lambda w: w['x0'])
                leftmost = grp[0]['x0']
                for w in grp:
                    is_ll = (abs(w['x0'] - leftmost) < 0.5)
                    ordered.append((w['text'], w['x0'], w['top'], is_ll,
                                    col_min_x, hw_thresh, pno))

    def unit_for(pno, top):
        """返回该页 top 位置所属的 unit 编号（None 表示无切换）。"""
        cur = None
        for btop, unum in unit_boundaries.get(pno, []):
            if top >= btop - LINE_TOL:
                cur = unum
            else:
                break
        return cur

    entries = []
    current_unit = None
    buf = []
    in_section = False

    def flush():
        nonlocal buf
        if buf and current_unit is not None:
            eng, pos, meaning = parse_entry(buf)
            if eng:
                entries.append({'unit': current_unit, 'english': eng,
                                'pos': pos, 'meaning': meaning})
        buf = []

    def buf_has_cjk(b):
        """判断 buf 中是否已出现中文（释义）。"""
        return any(re.search(r'[一-鿿]', x) for x in b)

    # 预扫描的 Unit 标题：{(pno, 标题top): unit_num}，用于精确触发切换。
    # 不用 top 容差归属：Unit 标题可能与上一单元最后一行同 top
    # （如左栏 o'clock 释义与右栏 Unit 3 标题同行高），容差会误切。
    unit_title_tops = {}
    for pno, bounds in unit_boundaries.items():
        for btop, unum in bounds:
            unit_title_tops[(pno, btop)] = unum

    N = len(ordered)
    i = 0
    while i < N:
        text, x0, top, is_ll, col_min_x, hw_thresh, pno = ordered[i]
        # Unit 切换：仅当当前 token 是 "Unit" 标题本身（top 精确匹配预扫描结果）
        tl = text.strip()
        if unit_re.match(tl) and (pno, top) in unit_title_tops:
            u = unit_title_tops[(pno, top)]
            if u != current_unit:
                flush()
                current_unit = u
                in_section = True
            i += 1
            continue
        if not in_section:
            i += 1
            continue
        t = text.strip()
        # 过滤页眉/页脚广告：注意 "关注" 可能是正常释义（如 "注意；关注"），
        # 因此只匹配组合关键词，避免误删
        if not t or t in JUNK or (t in ('and', 'in') and top < 130) \
           or re.search(r'微信公众号|关注微信|捷思课堂|获取更多学习资料', t) \
           or re.fullmatch(r'\d{1,3}', t):
            i += 1
            continue
        if is_headword_start(t, x0, is_ll, col_min_x, hw_thresh):
            # 跨行续词保护：若当前条目含未闭合括号（如 steal (stole, ...
            # 或 burn (burnt, ...），行首英文 token 是括号内的续词，
            # 不应作为新词头。全角（ ）也计入（教材排版可能混用）。
            if buf and (sum(x.count('(') + x.count('（') for x in buf)
                        > sum(x.count(')') + x.count('）') for x in buf)):
                buf.append(t)
                i += 1
                continue
            # 人名/专名跨行保护：同一行后续 token 中含中文，且当前条目尚无
            # 中文释义，说明这是释义续行（如 "Teresa" 下行的 "Lopez ... 特蕾莎"、
            # "J. K. Rowling" 释义行的 "J. K. 罗琳（"），不是新词头。
            nxt_has_cjk = False
            for j in range(i + 1, min(i + 6, N)):
                if abs(ordered[j][2] - top) > LINE_TOL:
                    break
                if re.search(r'[一-鿿]', ordered[j][0]):
                    nxt_has_cjk = True
                    break
            if nxt_has_cjk and not buf_has_cjk(buf):
                buf.append(t)
                i += 1
                continue
            flush()
            buf = [t]
        else:
            buf.append(t)
        i += 1
    flush()
    return entries


# ==================== Excel 输出 ====================
def write_excel(entries, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cnt = Counter(e['unit'] for e in entries)
    wb = Workbook()
    ws = wb.active
    ws.title = "词汇表"
    headers = ["单元", "英文单词/短语", "词性", "中文释义"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for e in entries:
        ws.append([f"Unit {e['unit']}", e['english'], e['pos'] or "其它", e['meaning']])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if c in (1, 3) else "left",
                                       wrap_text=(c == 4))
    ws.freeze_panes = "A2"
    widths = [10, 32, 16, 46]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("单元统计")
    ws2.append(["单元", "单词/短语数量"])
    for c in range(1, 3):
        cell = ws2.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for u in sorted(cnt):
        ws2.append([f"Unit {u}", cnt[u]])
    ws2.append(["合计", len(entries)])
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 3):
            ws2.cell(row=r, column=c).border = border
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 16

    wb.save(output_path)
    return len(entries)


# ==================== CLI 入口 ====================
def main():
    parser = argparse.ArgumentParser(description="英语教材 PDF 词汇表提取")
    parser.add_argument("pdf", help="PDF 文件路径")
    parser.add_argument("--pages", default=None, help="词汇表页码范围 1-based 闭区间，如 106-114")
    parser.add_argument("--cut", type=float, default=None, help="双栏切分 x 坐标")
    parser.add_argument("--top-skip", type=float, default=52.0, help="页眉跳过高度")
    parser.add_argument("--output", default="词汇表.xlsx", help="输出 Excel 路径")
    parser.add_argument("--unit-header", default="Unit", help="单元标题关键词")
    parser.add_argument("--json", default=None, help="同时输出 JSON 路径")
    parser.add_argument("--scan-only", action="store_true", help="仅扫描定位词汇表页码")
    parser.add_argument("--clean", action="store_true", help="提取完成后自动执行数据清洗，输出清洗后的Excel（单词|所属单元|词性|中文释义|例句）")
    parser.add_argument("--examples", action="store_true", help="清洗后自动为每个单词生成例句（从本教材正文提取，需配合--clean使用）")
    parser.add_argument("--llm-api-key", default=None, help="LLM API key（未匹配单词的例句生成兜底）")
    parser.add_argument("--llm-base-url", default=None, help="LLM API 基础地址")
    parser.add_argument("--llm-model", default=None, help="LLM 模型名")
    args = parser.parse_args()

    if not os.path.isfile(args.pdf):
        print(f"ERROR: PDF not found: {args.pdf}", file=sys.stderr)
        sys.exit(1)

    # 确定页码范围
    if args.pages:
        ps, pe = args.pages.split("-")
        page_start, page_end = int(ps), int(pe)
    else:
        print("正在扫描 PDF 定位词汇表...", file=sys.stderr)
        page_start, page_end = scan_for_vocab_pages(args.pdf, args.unit_header)
        if page_start is None:
            print("ERROR: 未找到词汇表页面。请用 --pages 手动指定页码范围。", file=sys.stderr)
            sys.exit(1)
        print(f"定位到词汇表: 第 {page_start}-{page_end} 页", file=sys.stderr)

    if args.scan_only:
        print(f"{page_start}-{page_end}")
        return

    # 确定栏切分：未指定时按页动态检测（不同页的双栏位置可能不同）
    cut = args.cut  # None 表示按页动态检测

    print(f"提取中: {args.pdf} 第{page_start}-{page_end}页 cut={'auto(按页)' if cut is None else f'{cut:.0f}'}", file=sys.stderr)
    entries = extract(args.pdf, page_start, page_end, cut, args.top_skip, args.unit_header)

    cnt = Counter(e['unit'] for e in entries)
    print(f"TOTAL: {len(entries)}", file=sys.stderr)
    print(f"PER UNIT: {dict(sorted(cnt.items()))}", file=sys.stderr)

    # 质检
    import re as _re
    bad_eng = [e for e in entries if _re.search(r'[一-鿿]', e['english']) or not e['english']]
    bad_pua = [e for e in entries if any(0xE000 <= ord(c) <= 0xF8FF for c in e['meaning'] + e['english'])]
    bad_par = [e for e in entries if e['english'].count('(') != e['english'].count(')')]
    print(f"质检: 英文异常={len(bad_eng)} 释义含PUA={len(bad_pua)} 括号不平衡={len(bad_par)}", file=sys.stderr)

    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(entries, f, ensure_ascii=False, indent=1)
        print(f"JSON: {args.json}", file=sys.stderr)

    n = write_excel(entries, args.output)
    print(f"Excel: {args.output} ({n} rows)", file=sys.stderr)

    # 自动数据清洗（--clean）
    if args.clean:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        if script_dir not in sys.path:
            sys.path.insert(0, script_dir)
        try:
            import clean_vocab
            cv_entries = [
                {"word": e["english"], "unit": e["unit"],
                 "pos": e["pos"], "meaning": e["meaning"]}
                for e in entries
            ]
            cleaned, stats = clean_vocab.clean_entries(cv_entries)
            
            # 自动生成例句（--examples，需配合--clean）
            if args.examples:
                print(f"\n=== 例句生成 ===", file=sys.stderr)
                ex_matched, ex_total, ex_unmatched = clean_vocab.generate_examples_for_entries(
                    cleaned, args.pdf, (page_start, page_end),
                    args.llm_api_key, args.llm_base_url, args.llm_model
                )
                print(f"  例句覆盖率: {ex_matched}/{ex_total} ({ex_matched*100//max(ex_total,1)}%)", file=sys.stderr)
                if ex_unmatched:
                    print(f"  未匹配单词: {len(ex_unmatched)} 个（可手动补充或配置--llm-api-key）", file=sys.stderr)
            
            check_passed, issues = clean_vocab.self_check(cleaned, len(entries))
            base, ext = os.path.splitext(args.output)
            clean_output = f"{base}_清洗后.xlsx"
            clean_vocab.write_excel(cleaned, clean_output)
            print(f"\n=== 数据清洗 ===", file=sys.stderr)
            clean_vocab.print_log(stats, len(entries), clean_output, check_passed, issues)
        except ImportError:
            print("WARNING: 未找到 clean_vocab.py，跳过清洗。请确保 clean_vocab.py 与本脚本在同一 scripts/ 目录下。", file=sys.stderr)


if __name__ == "__main__":
    main()
