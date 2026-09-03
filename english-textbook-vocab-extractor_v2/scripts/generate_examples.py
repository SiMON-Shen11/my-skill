# -*- coding: utf-8 -*-
"""
英语教材词汇表例句生成脚本。
优先从教材正文页提取包含目标单词的原句，提取不到的可用 LLM 生成兜底。

用法:
  python generate_examples.py <词汇表Excel> <教材PDF> [选项]

选项:
  --output PATH       输出带例句的 Excel 路径，默认 <输入名>_带例句.xlsx
  --llm-api-key KEY   LLM API key（用于未匹配单词的例句生成兜底）
  --llm-base-url URL  LLM API 基础地址（默认 OpenAI 兼容 https://api.openai.com/v1）
  --llm-model MODEL   LLM 模型名（默认 gpt-4o-mini）
  --max-sentences N   每个单词最多扫描的候选句子数（默认 50）
  --llm-only          全量使用LLM生成例句（跳过教材正文提取，需配合--llm-api-key）
  --json PATH         同时输出未匹配单词列表（JSON），便于后续人工/LLM补充
  --quiet             不输出处理日志
"""
import sys, os, re, json, argparse
from collections import defaultdict

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed. Run: pip install pdfplumber openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ==================== 不规则变化表 ====================
IRREGULAR_VERBS = {
    'be': ['am', 'is', 'are', 'was', 'were', 'been', 'being'],
    'have': ['has', 'had', 'having'],
    'do': ['does', 'did', 'done', 'doing'],
    'go': ['goes', 'went', 'gone', 'going'],
    'get': ['gets', 'got', 'gotten', 'getting'],
    'make': ['makes', 'made', 'making'],
    'take': ['takes', 'took', 'taken', 'taking'],
    'come': ['comes', 'came', 'coming'],
    'see': ['sees', 'saw', 'seen', 'seeing'],
    'know': ['knows', 'knew', 'known', 'knowing'],
    'think': ['thinks', 'thought', 'thinking'],
    'give': ['gives', 'gave', 'given', 'giving'],
    'find': ['finds', 'found', 'finding'],
    'tell': ['tells', 'told', 'telling'],
    'ask': ['asks', 'asked', 'asking'],
    'work': ['works', 'worked', 'working'],
    'seem': ['seems', 'seemed', 'seeming'],
    'feel': ['feels', 'felt', 'feeling'],
    'try': ['tries', 'tried', 'trying'],
    'leave': ['leaves', 'left', 'leaving'],
    'call': ['calls', 'called', 'calling'],
    'need': ['needs', 'needed', 'needing'],
    'become': ['becomes', 'became', 'become', 'becoming'],
    'put': ['puts', 'put', 'putting'],
    'mean': ['means', 'meant', 'meaning'],
    'keep': ['keeps', 'kept', 'keeping'],
    'let': ['lets', 'let', 'letting'],
    'begin': ['begins', 'began', 'begun', 'beginning'],
    'help': ['helps', 'helped', 'helping'],
    'show': ['shows', 'showed', 'shown', 'showing'],
    'hear': ['hears', 'heard', 'hearing'],
    'play': ['plays', 'played', 'playing'],
    'run': ['runs', 'ran', 'running'],
    'move': ['moves', 'moved', 'moving'],
    'live': ['lives', 'lived', 'living'],
    'believe': ['believes', 'believed', 'believing'],
    'bring': ['brings', 'brought', 'bringing'],
    'happen': ['happens', 'happened', 'happening'],
    'write': ['writes', 'wrote', 'written', 'writing'],
    'sit': ['sits', 'sat', 'sitting'],
    'stand': ['stands', 'stood', 'standing'],
    'lose': ['loses', 'lost', 'losing'],
    'pay': ['pays', 'paid', 'paying'],
    'meet': ['meets', 'met', 'meeting'],
    'include': ['includes', 'included', 'including'],
    'continue': ['continues', 'continued', 'continuing'],
    'set': ['sets', 'set', 'setting'],
    'learn': ['learns', 'learned', 'learnt', 'learning'],
    'change': ['changes', 'changed', 'changing'],
    'lead': ['leads', 'led', 'leading'],
    'understand': ['understands', 'understood', 'understanding'],
    'watch': ['watches', 'watched', 'watching'],
    'follow': ['follows', 'followed', 'following'],
    'stop': ['stops', 'stopped', 'stopping'],
    'create': ['creates', 'created', 'creating'],
    'speak': ['speaks', 'spoke', 'spoken', 'speaking'],
    'read': ['reads', 'read', 'reading'],
    'allow': ['allows', 'allowed', 'allowing'],
    'add': ['adds', 'added', 'adding'],
    'spend': ['spends', 'spent', 'spending'],
    'grow': ['grows', 'grew', 'grown', 'growing'],
    'open': ['opens', 'opened', 'opening'],
    'walk': ['walks', 'walked', 'walking'],
    'win': ['wins', 'won', 'winning'],
    'offer': ['offers', 'offered', 'offering'],
    'remember': ['remembers', 'remembered', 'remembering'],
    'love': ['loves', 'loved', 'loving'],
    'consider': ['considers', 'considered', 'considering'],
    'appear': ['appears', 'appeared', 'appearing'],
    'buy': ['buys', 'bought', 'buying'],
    'wait': ['waits', 'waited', 'waiting'],
    'serve': ['serves', 'served', 'serving'],
    'die': ['dies', 'died', 'dying'],
    'send': ['sends', 'sent', 'sending'],
    'expect': ['expects', 'expected', 'expecting'],
    'build': ['builds', 'built', 'building'],
    'stay': ['stays', 'stayed', 'staying'],
    'fall': ['falls', 'fell', 'fallen', 'falling'],
    'cut': ['cuts', 'cut', 'cutting'],
    'reach': ['reaches', 'reached', 'reaching'],
    'kill': ['kills', 'killed', 'killing'],
    'remain': ['remains', 'remained', 'remaining'],
    'suggest': ['suggests', 'suggested', 'suggesting'],
    'raise': ['raises', 'raised', 'raising'],
    'pass': ['passes', 'passed', 'passing'],
    'sell': ['sells', 'sold', 'selling'],
    'require': ['requires', 'required', 'requiring'],
    'report': ['reports', 'reported', 'reporting'],
    'decide': ['decides', 'decided', 'deciding'],
    'pull': ['pulls', 'pulled', 'pulling'],
    'develop': ['develops', 'developed', 'developing'],
    'carry': ['carries', 'carried', 'carrying'],
    'thank': ['thanks', 'thanked', 'thanking'],
    'receive': ['receives', 'received', 'receiving'],
    'agree': ['agrees', 'agreed', 'agreeing'],
    'support': ['supports', 'supported', 'supporting'],
    'hit': ['hits', 'hit', 'hitting'],
    'produce': ['produces', 'produced', 'producing'],
    'eat': ['eats', 'ate', 'eaten', 'eating'],
    'cover': ['covers', 'covered', 'covering'],
    'catch': ['catches', 'caught', 'catching'],
    'draw': ['draws', 'drew', 'drawn', 'drawing'],
    'choose': ['chooses', 'chose', 'chosen', 'choosing'],
    'fly': ['flies', 'flew', 'flown', 'flying'],
    'feed': ['feeds', 'fed', 'feeding'],
    'steal': ['steals', 'stole', 'stolen', 'stealing'],
    'wear': ['wears', 'wore', 'worn', 'wearing'],
    'forget': ['forgets', 'forgot', 'forgotten', 'forgetting'],
    'drive': ['drives', 'drove', 'driven', 'driving'],
    'ride': ['rides', 'rode', 'ridden', 'riding'],
    'rise': ['rises', 'rose', 'risen', 'rising'],
    'shake': ['shakes', 'shook', 'shaken', 'shaking'],
    'sing': ['sings', 'sang', 'sung', 'singing'],
    'sleep': ['sleeps', 'slept', 'sleeping'],
    'swim': ['swims', 'swam', 'swum', 'swimming'],
    'teach': ['teaches', 'taught', 'teaching'],
    'throw': ['throws', 'threw', 'thrown', 'throwing'],
    'wake': ['wakes', 'woke', 'woken', 'waking'],
    'fight': ['fights', 'fought', 'fighting'],
    'find': ['finds', 'found', 'finding'],
    'hang': ['hangs', 'hung', 'hanged', 'hanging'],
    'hide': ['hides', 'hid', 'hidden', 'hiding'],
    'hold': ['holds', 'held', 'holding'],
    'lay': ['lays', 'laid', 'laying'],
    'lead': ['leads', 'led', 'leading'],
    'lend': ['lends', 'lent', 'lending'],
    'light': ['lights', 'lit', 'lighted', 'lighting'],
    'meet': ['meets', 'met', 'meeting'],
    'prove': ['proves', 'proved', 'proven', 'proving'],
    'ring': ['rings', 'rang', 'rung', 'ringing'],
    'rise': ['rises', 'rose', 'risen', 'rising'],
    'sew': ['sews', 'sewed', 'sewn', 'sewing'],
    'shake': ['shakes', 'shook', 'shaken', 'shaking'],
    'shoot': ['shoots', 'shot', 'shooting'],
    'show': ['shows', 'showed', 'shown', 'showing'],
    'shut': ['shuts', 'shut', 'shutting'],
    'sing': ['sings', 'sang', 'sung', 'singing'],
    'sink': ['sinks', 'sank', 'sunk', 'sinking'],
    'sit': ['sits', 'sat', 'sitting'],
    'sleep': ['sleeps', 'slept', 'sleeping'],
    'slide': ['slides', 'slid', 'sliding'],
    'speak': ['speaks', 'spoke', 'spoken', 'speaking'],
    'speed': ['speeds', 'sped', 'speeded', 'speeding'],
    'spell': ['spells', 'spelt', 'spelled', 'spelling'],
    'spend': ['spends', 'spent', 'spending'],
    'spill': ['spills', 'spilt', 'spilled', 'spilling'],
    'spin': ['spins', 'spun', 'spinning'],
    'spit': ['spits', 'spat', 'spitting'],
    'split': ['splits', 'split', 'splitting'],
    'spoil': ['spoils', 'spoilt', 'spoiled', 'spoiling'],
    'spread': ['spreads', 'spread', 'spreading'],
    'spring': ['springs', 'sprang', 'sprung', 'springing'],
    'stand': ['stands', 'stood', 'standing'],
    'steal': ['steals', 'stole', 'stolen', 'stealing'],
    'stick': ['sticks', 'stuck', 'sticking'],
    'sting': ['stings', 'stung', 'stinging'],
    'stink': ['stinks', 'stank', 'stunk', 'stinking'],
    'strike': ['strikes', 'struck', 'struck', 'striking'],
    'string': ['strings', 'strung', 'stringing'],
    'strive': ['strives', 'strove', 'striven', 'striving'],
    'swear': ['swears', 'swore', 'sworn', 'swearing'],
    'sweep': ['sweeps', 'swept', 'sweeping'],
    'swell': ['swells', 'swelled', 'swollen', 'swelling'],
    'swim': ['swims', 'swam', 'swum', 'swimming'],
    'swing': ['swings', 'swung', 'swinging'],
    'take': ['takes', 'took', 'taken', 'taking'],
    'teach': ['teaches', 'taught', 'teaching'],
    'tear': ['tears', 'tore', 'torn', 'tearing'],
    'tell': ['tells', 'told', 'telling'],
    'think': ['thinks', 'thought', 'thinking'],
    'throw': ['throws', 'threw', 'thrown', 'throwing'],
    'thrust': ['thrusts', 'thrust', 'thrusting'],
    'understand': ['understands', 'understood', 'understanding'],
    'uphold': ['upholds', 'upheld', 'upholding'],
    'upset': ['upsets', 'upset', 'upsetting'],
    'wake': ['wakes', 'woke', 'woken', 'waking'],
    'wear': ['wears', 'wore', 'worn', 'wearing'],
    'weave': ['weaves', 'wove', 'woven', 'weaving'],
    'wed': ['weds', 'wed', 'wedded', 'wedding'],
    'weep': ['weeps', 'wept', 'weeping'],
    'wet': ['wets', 'wet', 'wetted', 'wetting'],
    'win': ['wins', 'won', 'winning'],
    'wind': ['winds', 'wound', 'winding'],
    'withdraw': ['withdraws', 'withdrew', 'withdrawn', 'withdrawing'],
    'withhold': ['withholds', 'withheld', 'withholding'],
    'withstand': ['withstands', 'withstood', 'withstanding'],
    'wrap': ['wraps', 'wrapped', 'wrapping'],
    'wring': ['wrings', 'wrung', 'wringing'],
    'write': ['writes', 'wrote', 'written', 'writing'],
}

IRREGULAR_NOUNS = {
    'child': ['children'],
    'foot': ['feet'],
    'tooth': ['teeth'],
    'man': ['men'],
    'woman': ['women'],
    'mouse': ['mice'],
    'goose': ['geese'],
    'leaf': ['leaves'],
    'life': ['lives'],
    'knife': ['knives'],
    'wife': ['wives'],
    'half': ['halves'],
    'shelf': ['shelves'],
    'wolf': ['wolves'],
    'thief': ['thieves'],
    'self': ['selves'],
    'person': ['people', 'persons'],
    'ox': ['oxen'],
    'crisis': ['crises'],
    'analysis': ['analyses'],
    'thesis': ['theses'],
    'criterion': ['criteria'],
    'phenomenon': ['phenomena'],
    'bacterium': ['bacteria'],
    'datum': ['data'],
    'medium': ['media', 'mediums'],
    'index': ['indices', 'indexes'],
    'matrix': ['matrices', 'matrixes'],
    'formula': ['formulae', 'formulas'],
    'vita': ['vitae'],
    'alumnus': ['alumni'],
    'alumna': ['alumnae'],
    'cactus': ['cacti', 'cactuses'],
    'fungus': ['fungi', 'funguses'],
    'nucleus': ['nuclei', 'nucleuses'],
    'syllabus': ['syllabi', 'syllabuses'],
    'terminus': ['termini', 'terminuses'],
    'erratum': ['errata'],
    'addendum': ['addenda'],
    'memorandum': ['memoranda', 'memorandums'],
    'ovum': ['ova'],
    'stratum': ['strata'],
    'focus': ['foci', 'focuses'],
    'fungus': ['fungi', 'funguses'],
    'genus': ['genera'],
    'species': ['species'],
    'series': ['series'],
    'aircraft': ['aircraft'],
    'fish': ['fish', 'fishes'],
    'sheep': ['sheep'],
    'deer': ['deer'],
    'moose': ['moose'],
    'swine': ['swine'],
    'bison': ['bison'],
    'salmon': ['salmon'],
    'trout': ['trout'],
    'shrimp': ['shrimp', 'shrimps'],
    'means': ['means'],
    'headquarters': ['headquarters'],
    'works': ['works'],
    'crossroads': ['crossroads'],
    'barracks': ['barracks'],
    'gallows': ['gallows'],
    'clothes': ['clothes'],
    'odds': ['odds'],
    'remains': ['remains'],
    'thanks': ['thanks'],
    'wages': ['wages'],
    'scissors': ['scissors'],
    'trousers': ['trousers'],
    'pants': ['pants'],
    'shorts': ['shorts'],
    'glasses': ['glasses'],
    'jeans': ['jeans'],
    'pyjamas': ['pyjamas'],
    'pajamas': ['pajamas'],
    'binoculars': ['binoculars'],
    'pliers': ['pliers'],
    'tongs': ['tongs'],
    'shears': ['shears'],
    'spectacles': ['spectacles'],
}


# ==================== 词形变化生成 ====================
def generate_word_forms(word):
    """生成单词的所有可能形式（原形+规则变化+不规则变化）。"""
    word = word.lower().strip()
    forms = {word}
    
    # 不规则动词
    if word in IRREGULAR_VERBS:
        forms.update(IRREGULAR_VERBS[word])
    
    # 不规则名词
    if word in IRREGULAR_NOUNS:
        forms.update(IRREGULAR_NOUNS[word])
    
    # 规则动词变化
    # 第三人称单数
    if word.endswith('s') or word.endswith('x') or word.endswith('z') or word.endswith('ch') or word.endswith('sh'):
        forms.add(word + 'es')
    elif word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
        forms.add(word[:-1] + 'ies')
    else:
        forms.add(word + 's')
    
    # 过去式/过去分词
    if word.endswith('e'):
        forms.add(word + 'd')
        forms.add(word + 'ing')
    elif word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
        forms.add(word[:-1] + 'ied')
        forms.add(word[:-1] + 'ying')
    elif len(word) >= 3 and word[-1] not in 'aeiou' and word[-2] in 'aeiou' and word[-3] not in 'aeiou':
        # 重读闭音节双写
        forms.add(word + word[-1] + 'ed')
        forms.add(word + word[-1] + 'ing')
    else:
        forms.add(word + 'ed')
        forms.add(word + 'ing')
    
    # 名词复数
    if word.endswith('s') or word.endswith('x') or word.endswith('z') or word.endswith('ch') or word.endswith('sh'):
        forms.add(word + 'es')
    elif word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
        forms.add(word[:-1] + 'ies')
    elif word.endswith('f'):
        forms.add(word[:-1] + 'ves')
    elif word.endswith('fe'):
        forms.add(word[:-2] + 'ves')
    else:
        forms.add(word + 's')
    
    # 形容词比较级/最高级
    if len(word) <= 6:
        if word.endswith('e'):
            forms.add(word + 'r')
            forms.add(word + 'st')
        elif word.endswith('y') and len(word) > 1 and word[-2] not in 'aeiou':
            forms.add(word[:-1] + 'ier')
            forms.add(word[:-1] + 'iest')
        elif len(word) >= 3 and word[-1] not in 'aeiou' and word[-2] in 'aeiou' and word[-3] not in 'aeiou':
            forms.add(word + word[-1] + 'er')
            forms.add(word + word[-1] + 'est')
        else:
            forms.add(word + 'er')
            forms.add(word + 'est')
    
    return forms


def generate_split_forms(word):
    """生成单词可能的PDF拆分形式（在任意位置插入空格）。
    用于匹配PDF中被字间距拆分的单词，如 diff erent, fi nd, Th ey。
    """
    word = word.lower().strip()
    if len(word) < 4:
        return {word}
    forms = {word}
    # 在第2到倒数第2个字符之间插入空格（避免单字母拆分）
    for i in range(2, len(word) - 1):
        split = word[:i] + ' ' + word[i:]
        forms.add(split)
        # 也支持首字母大写形式
        forms.add(split.capitalize())
    forms.add(word.capitalize())
    return forms


def extract_phrase_keywords(phrase):
    """从短语中提取核心实词（去掉占位符、介词、冠词等）。"""
    # 去掉括号内容和占位符
    cleaned = re.sub(r'\([^)]*\)', '', phrase)
    cleaned = cleaned.replace('...', '').replace('⋯', '').replace('…', '')
    # 分割成单词
    words = re.findall(r"[A-Za-z']+", cleaned.lower())
    # 去掉虚词
    stopwords = {'a', 'an', 'the', 'to', 'of', 'in', 'on', 'at', 'for', 'with', 'from', 'by', 'about', 'as', 'into', 'onto', 'upon', 'and', 'or', 'but', 'so', 'if', 'because', 'while', 'when', 'where', 'which', 'who', 'whom', 'whose', 'that', 'this', 'these', 'those', 'it', 'its', 'i', 'me', 'my', 'we', 'us', 'our', 'you', 'your', 'he', 'him', 'his', 'she', 'her', 'hers', 'they', 'them', 'their', 'be', 'is', 'am', 'are', 'was', 'were', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'shall', 'should', 'can', 'could', 'may', 'might', 'must', 'ought', 'dare', 'need', 'used', 'up', 'down', 'out', 'off', 'over', 'under', 'again', 'further', 'then', 'once', 'here', 'there', 'when', 'where', 'why', 'how', 'all', 'any', 'both', 'each', 'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'than', 'too', 'very', 'just', 'also', 'now', 'sb', 'sth', 'someone', 'something', 'somebody'}
    keywords = [w for w in words if w not in stopwords and len(w) > 1]
    return keywords if keywords else words


# ==================== 正文提取 ====================
def is_word_list_line(line):
    """判断一行是否是单词列表行（逗号分隔的短词，如单元介绍页的词汇列表）。"""
    # 统计逗号分隔的片段
    parts = [p.strip() for p in line.split(',') if p.strip()]
    if len(parts) >= 4:
        # 大部分片段是短词（<=2个单词）
        short_count = sum(1 for p in parts if len(p.split()) <= 2)
        if short_count >= len(parts) * 0.7:
            return True
    return False


# 初中教材常见PDF拆分单词修复表（键：小写拆分形式，值：正确形式）
SPLIT_WORD_FIX = {
    'th ey': 'they', 'th eir': 'their', 'th em': 'them', 'th en': 'then',
    'th ere': 'there', 'th ese': 'these', 'th ose': 'those', 'th ough': 'though',
    'th rough': 'through', 'th row': 'throw', 'th ick': 'thick', 'th in': 'thin',
    'th ing': 'thing', 'th ink': 'think', 'th ird': 'third', 'th irty': 'thirty',
    'th is': 'this', 'th at': 'that', 'th an': 'than', 'th ank': 'thank',
    'th eatre': 'theatre', 'th eater': 'theater', 'th eft': 'theft',
    'th eme': 'theme', 'th eory': 'theory', 'th erefore': 'therefore',
    'th erapy': 'therapy', 'th ermal': 'thermal', 'th esis': 'thesis',
    'th ick': 'thick', 'th ief': 'thief', 'th igh': 'thigh', 'th imble': 'thimble',
    'th irst': 'thirst', 'th irteen': 'thirteen', 'th irty': 'thirty',
    'th orough': 'thorough', 'th read': 'thread', 'th reat': 'threat',
    'th ree': 'three', 'th resh': 'thresh', 'th ride': 'thride',
    'th rift': 'thrift', 'th rill': 'thrill', 'th rive': 'thrive',
    'th roat': 'throat', 'th rong': 'throng', 'th rone': 'throne',
    'th uder': 'thunder', 'th umb': 'thumb', 'th us': 'thus',
    'diff erent': 'different', 'diff icult': 'difficult', 'diff erence': 'difference',
    'diff use': 'diffuse', 'diff ident': 'diffident',
    'fi nd': 'find', 'fi nish': 'finish', 'fi rst': 'first', 'fi sh': 'fish',
    'fi gure': 'figure', 'fi eld': 'field', 'fi ght': 'fight', 'fi lled': 'filled',
    'fi lm': 'film', 'fi lter': 'filter', 'fi nal': 'final', 'fi nance': 'finance',
    'fi nger': 'finger', 'fi re': 'fire', 'fi rm': 'firm', 'fi x': 'fix',
    'fi t': 'fit', 'fi ve': 'five', 'fi xed': 'fixed',
    'fl y': 'fly', 'fl ow': 'flow', 'fl oor': 'floor', 'fl ower': 'flower',
    'fl at': 'flat', 'fl avor': 'flavor', 'fl eet': 'fleet', 'fl esh': 'flesh',
    'fl ight': 'flight', 'fl ood': 'flood', 'fl our': 'flour', 'fl uent': 'fluent',
    'fl uid': 'fluid', 'fl ush': 'flush', 'fl ying': 'flying',
    'aft er': 'after', 'aft ernoon': 'afternoon', 'aft erwards': 'afterwards',
    'oft en': 'often', 'bett er': 'better', 'writ er': 'writer', 'read er': 'reader',
    'speak er': 'speaker', 'teach er': 'teacher', 'work er': 'worker',
    'play er': 'player', 'sing er': 'singer', 'danc er': 'dancer',
    'run er': 'runner', 'swim er': 'swimmer', 'win er': 'winner',
    'impr ove': 'improve', 'impr ovement': 'improvement',
    'att ention': 'attention', 'att ract': 'attract', 'att raction': 'attraction',
    'att empt': 'attempt', 'att end': 'attend', 'att itude': 'attitude',
    'conn ect': 'connect', 'conn ection': 'connection',
    'expr ession': 'expression', 'expr ess': 'express',
    'sent ence': 'sentence', 'stud y': 'study', 'stud ent': 'student',
    'wat ch': 'watch', 'wat er': 'water', 'litt le': 'little',
    'enou gh': 'enough', 'thr ough': 'through', 'anot her': 'another',
    'impo rtant': 'important', 'impo rtance': 'importance',
    'info rmation': 'information', 'inte resting': 'interesting',
    'inte rnational': 'international', 'inte rnet': 'internet',
    'pron ounce': 'pronounce', 'pron unciation': 'pronunciation',
    'prac tice': 'practice', 'prac tical': 'practical',
    'lang uage': 'language', 'wond erful': 'wonderful',
    'bea utiful': 'beautiful', 'care fully': 'carefully',
    'success ful': 'successful', 'espe cially': 'especially',
    'actu ally': 'actually', 'cert ainly': 'certainly',
    'proba bly': 'probably', 'comp letely': 'completely',
    'deve lopment': 'development', 'enviro nment': 'environment',
    'expe rience': 'experience', 'expe riment': 'experiment',
    'gove rnment': 'government', 'histo rical': 'historical',
    'imme diately': 'immediately', 'lite rature': 'literature',
    'mathe matics': 'mathematics', 'mech anical': 'mechanical',
    'medi cine': 'medicine', 'natu ral': 'natural',
    'nece ssary': 'necessary', 'occa sionally': 'occasionally',
    'offi cial': 'official', 'offi ce': 'office',
    'ope ration': 'operation', 'oppo rtunity': 'opportunity',
    'orga nization': 'organization', 'part icular': 'particular',
    'perc entage': 'percentage', 'perh aps': 'perhaps',
    'phy sical': 'physical', 'poli tical': 'political',
    'popu lation': 'population', 'posi tion': 'position',
    'poss ible': 'possible', 'pre fer': 'prefer',
    'pre sent': 'present', 'pre sident': 'president',
    'pre ssure': 'pressure', 'pre tty': 'pretty',
    'pre vent': 'prevent', 'pri mary': 'primary',
    'pri vate': 'private', 'pro blem': 'problem',
    'pro cess': 'process', 'pro duce': 'produce',
    'pro duct': 'product', 'pro duction': 'production',
    'pro fessor': 'professor', 'pro gram': 'program',
    'pro gress': 'progress', 'pro ject': 'project',
    'pro mise': 'promise', 'pro nounce': 'pronounce',
    'pro per': 'proper', 'pro portion': 'proportion',
    'pro posal': 'proposal', 'pro pose': 'propose',
    'pro spect': 'prospect', 'pro tect': 'protect',
    'pro tein': 'protein', 'pro test': 'protest',
    'pro vide': 'provide', 'pro vince': 'province',
    'pub lic': 'public', 'pub lish': 'publish',
    'pur pose': 'purpose', 'qual ity': 'quality',
    'quan tity': 'quantity', 'quar ter': 'quarter',
    'ques tion': 'question', 'reali ze': 'realize',
    'reco gnize': 'recognize', 'reco rd': 'record',
    'reduc e': 'reduce', 'refl ect': 'reflect',
    'refu se': 'refuse', 'regi on': 'region',
    'regu lar': 'regular', 'rela tion': 'relation',
    'rela tionship': 'relationship', 'rela tive': 'relative',
    'rele ase': 'release', 'rele vant': 'relevant',
    'reli gion': 'religion', 'rema in': 'remain',
    'reme mber': 'remember', 'remi nd': 'remind',
    'remo te': 'remote', 'remo ve': 'remove',
    'repa ir': 'repair', 'repe at': 'repeat',
    'repl ace': 'replace', 'repl y': 'reply',
    'repo rt': 'report', 'repre sent': 'represent',
    'repu blic': 'republic', 'reque st': 'request',
    'requ ire': 'require', 'resea rch': 'research',
    'reser ve': 'reserve', 'resi dent': 'resident',
    'resist': 'resist', 'resou rce': 'resource',
    'respe ct': 'respect', 'respo nd': 'respond',
    'respo nse': 'response', 'resta urant': 'restaurant',
    'resto re': 'restore', 'resu lt': 'result',
    'retire': 'retire', 'retu rn': 'return',
    'reveal': 'reveal', 'review': 'review',
    'revol ution': 'revolution', 'rewa rd': 'reward',
    'satis fy': 'satisfy', 'sched ule': 'schedule',
    'scien ce': 'science', 'scree n': 'screen',
    'searc h': 'search', 'seaso n': 'season',
    'seco nd': 'second', 'secre t': 'secret',
    'secti on': 'section', 'secu re': 'secure',
    'senti ment': 'sentiment', 'separ ate': 'separate',
    'serie s': 'series', 'serio us': 'serious',
    'serv ant': 'servant', 'serv ice': 'service',
    'sessi on': 'session', 'sett le': 'settle',
    'sever al': 'several', 'shad ow': 'shadow',
    'shake': 'shake', 'shall': 'shall',
    'shap e': 'shape', 'share': 'share',
    'sharp': 'sharp', 'sheet': 'sheet',
    'shelf': 'shelf', 'shell': 'shell',
    'shel ter': 'shelter', 'shift': 'shift',
    'shine': 'shine', 'ship': 'ship',
    'shirt': 'shirt', 'shock': 'shock',
    'shoe': 'shoe', 'shoot': 'shoot',
    'shop': 'shop', 'shore': 'shore',
    'short': 'short', 'shot': 'shot',
    'should': 'should', 'shoulder': 'shoulder',
    'shout': 'shout', 'show': 'show',
    'shower': 'shower', 'shut': 'shut',
    'sick': 'sick', 'side': 'side',
    'sight': 'sight', 'sign': 'sign',
    'signal': 'signal', 'silence': 'silence',
    'silent': 'silent', 'silk': 'silk',
    'silly': 'silly', 'silver': 'silver',
    'similar': 'similar', 'simple': 'simple',
    'since': 'since', 'sing': 'sing',
    'singer': 'singer', 'single': 'single',
    'sink': 'sink', 'sir': 'sir',
    'sister': 'sister', 'sit': 'sit',
    'site': 'site', 'situation': 'situation',
    'six': 'six', 'size': 'size',
    'skate': 'skate', 'skill': 'skill',
    'skin': 'skin', 'skirt': 'skirt',
    'sky': 'sky', 'slave': 'slave',
    'sleep': 'sleep', 'slow': 'slow',
    'small': 'small', 'smart': 'smart',
    'smell': 'smell', 'smile': 'smile',
    'smoke': 'smoke', 'smooth': 'smooth',
    'snake': 'snake', 'snow': 'snow',
    'so': 'so', 'soap': 'soap',
    'social': 'social', 'society': 'society',
    'sock': 'sock', 'sofa': 'sofa',
    'soft': 'soft', 'software': 'software',
    'soil': 'soil', 'solar': 'solar',
    'soldier': 'soldier', 'solid': 'solid',
    'solution': 'solution', 'solve': 'solve',
    'some': 'some', 'somebody': 'somebody',
    'someone': 'someone', 'something': 'something',
    'sometimes': 'sometimes', 'son': 'son',
    'song': 'song', 'soon': 'soon',
    'sort': 'sort', 'soul': 'soul',
    'sound': 'sound', 'soup': 'soup',
    'source': 'source', 'south': 'south',
    'southern': 'southern', 'space': 'space',
    'speak': 'speak', 'speaker': 'speaker',
    'special': 'special', 'species': 'species',
    'specific': 'specific', 'speech': 'speech',
    'speed': 'speed', 'spell': 'spell',
    'spend': 'spend', 'spin': 'spin',
    'spirit': 'spirit', 'spite': 'spite',
    'sport': 'sport', 'spot': 'spot',
    'spread': 'spread', 'spring': 'spring',
    'square': 'square', 'staff': 'staff',
    'stage': 'stage', 'stair': 'stair',
    'stamp': 'stamp', 'stand': 'stand',
    'standard': 'standard', 'star': 'star',
    'stare': 'stare', 'start': 'start',
    'state': 'state', 'statement': 'statement',
    'station': 'station', 'stay': 'stay',
    'steady': 'steady', 'steal': 'steal',
    'steam': 'steam', 'steel': 'steel',
    'step': 'step', 'stick': 'stick',
    'still': 'still', 'stomach': 'stomach',
    'stone': 'stone', 'stop': 'stop',
    'store': 'store', 'storm': 'storm',
    'story': 'story', 'stove': 'stove',
    'straight': 'straight', 'strange': 'strange',
    'stranger': 'stranger', 'straw': 'straw',
    'stream': 'stream', 'street': 'street',
    'strength': 'strength', 'stress': 'stress',
    'stretch': 'stretch', 'strict': 'strict',
    'strike': 'strike', 'string': 'string',
    'strip': 'strip', 'stroke': 'stroke',
    'strong': 'strong', 'structure': 'structure',
    'struggle': 'struggle', 'student': 'student',
    'study': 'study', 'stuff': 'stuff',
    'stupid': 'stupid', 'style': 'style',
    'subject': 'subject', 'submit': 'submit',
    'subscribe': 'subscribe', 'substance': 'substance',
    'succeed': 'succeed', 'success': 'success',
    'successful': 'successful', 'such': 'such',
    'sudden': 'sudden', 'suffer': 'suffer',
    'suggest': 'suggest', 'suggestion': 'suggestion',
    'suit': 'suit', 'suitable': 'suitable',
    'sum': 'sum', 'summer': 'summer',
    'sun': 'sun', 'sunday': 'sunday',
    'sunny': 'sunny', 'super': 'super',
    'supply': 'supply', 'support': 'support',
    'suppose': 'suppose', 'sure': 'sure',
    'surface': 'surface', 'surprise': 'surprise',
    'surround': 'surround', 'survey': 'survey',
    'survive': 'survive', 'suspect': 'suspect',
    'swallow': 'swallow', 'swear': 'swear',
    'sweat': 'sweat', 'sweep': 'sweep',
    'sweet': 'sweet', 'swim': 'swim',
    'swing': 'swing', 'switch': 'switch',
    'sword': 'sword', 'symbol': 'symbol',
    'system': 'system',
}


def post_process_example(s):
    """例句后处理：修复PDF拆分单词、首字母大写、清理多余空格。"""
    if not s:
        return s
    s = s.strip()
    # 修复拆分单词（大小写不敏感替换）
    s_lower = s.lower()
    for split_form, correct_form in SPLIT_WORD_FIX.items():
        if split_form in s_lower:
            # 保留原大小写的首字母
            idx = s_lower.find(split_form)
            original = s[idx:idx+len(split_form)]
            if original[0].isupper():
                replacement = correct_form.capitalize()
            else:
                replacement = correct_form
            s = s[:idx] + replacement + s[idx+len(split_form):]
            s_lower = s.lower()
    # 清理多余空格
    s = re.sub(r'\s+', ' ', s).strip()
    # 首字母大写（如果第一个字符是字母）
    if s and s[0].isalpha() and s[0].islower():
        s = s[0].upper() + s[1:]
    return s


def is_high_quality_example(s):
    """评估例句质量，返回 True 表示高质量（不需要LLM替换）。"""
    if not s:
        return False
    s = s.strip()
    # 含练习题关键词
    exercise_keywords = r'\b(Listen|Read|Match|Circle|Write|Complete|Fill|Choose|Correct|Translate|Role-play|Section|Grammar|Focus|Unit|Lesson|Exercise|Question|Answer|Conversation|Practice|Practice|Check|Number|Order|Repeat|Say|Ask|Tell|Look|Find|Draw|Color|Colour|Underline|Highlight)\b'
    if re.search(exercise_keywords, s, re.IGNORECASE):
        # 但如果这些词在句子中间且是正常用法（如 "I read a book"），不排除
        # 只排除以这些词开头或含数字+字母编号的
        if re.match(r'^(' + exercise_keywords + r')', s, re.IGNORECASE):
            return False
        if re.search(r'\b\d+[a-z]\b', s):
            return False
    # 含数字+字母编号（如 1c, 2a）
    if re.search(r'\b\d+[a-z]\b', s):
        return False
    # 含 "e.g." 或 "i.e." 且后面是单词列表（通常是练习题示例）
    if re.search(r'e\.g\.\s+[a-z]', s, re.IGNORECASE):
        return False
    # 含 "Q:" 或 "A:" 对话标记
    if re.search(r'\b[QA]:', s):
        return False
    # 单词列表行（多个短词用空格或逗号分隔，无动词）
    words = s.split()
    if len(words) >= 4:
        has_verb = any(re.search(r'(ing|ed|s)$', w.lower()) for w in words if len(w) > 3)
        if not has_verb and ',' not in s:
            # 可能是单词列表
            short_words = sum(1 for w in words if len(w) <= 4)
            if short_words >= len(words) * 0.7:
                return False
    # 首字母大写
    if s and s[0].isalpha() and not s[0].isupper():
        return False
    # 以标点结尾
    if not re.search(r'[.?!"\'\)]$', s):
        return False
    return True


def is_valid_sentence(s):
    """判断是否是有效的英文句子。"""
    s = s.strip()
    if not s:
        return False
    # 排除含特殊符号的句子（项目符号 ➊➋ 等、注册商标 ® 等）
    if re.search(r'[\u2460-\u24FF\u00AE\u2122]', s):
        return False
    # 排除含 "Page \d+" 页码引用的句子
    if re.search(r'Page\s+\d+', s, re.IGNORECASE):
        return False
    # 排除目录关键词
    if re.search(r'Units\s+Topics\s+Functions|Target\s+Language\s+Vocabulary', s):
        return False
    # 排除练习题指令（以数字+字母开头，如 2a, 1b, 3c）
    if re.match(r'^\d+[a-z]\s', s):
        return False
    # 中文占比不超过40%（排除中文翻译行）
    cjk_count = len(re.findall(r'[\u4e00-\u9fff]', s))
    if cjk_count > len(s) * 0.4:
        return False
    # 单词数在合理范围
    word_count = len(s.split())
    if word_count < 4 or word_count > 40:
        return False
    # 必须包含至少一个英文单词（3个字母以上）
    if not re.search(r'[a-zA-Z]{3,}', s):
        return False
    # 排除含词汇表页码引用（p.101）的句子
    if re.search(r'p\.\d+', s):
        return False
    # 排除含音标模式（/.../ 中含音标特征字符）的句子
    if re.search(r'/[a-zA-Z@:(){}\[\]ɜəɪʊɔæʌθðʃʒŋ]+/', s):
        return False
    # 排除含 PUA 字符的句子（七年级教材音标用PUA）
    if re.search(r'[\ue000-\uf8ff]', s):
        return False
    # 句子必须以 .?! 或引号/括号结尾
    if not re.search(r'[.?!\"\'\)]$', s):
        return False
    # 逗号分隔的短词过多（单词列表行）
    comma_parts = [p.strip() for p in s.split(',') if p.strip()]
    if len(comma_parts) >= 5:
        short_count = sum(1 for p in comma_parts if len(p.split()) <= 2)
        if short_count >= len(comma_parts) * 0.7:
            return False
    return True


def extract_body_text(pdf_path, vocab_pages=None, skip_front_pages=8):
    """从教材PDF提取正文页文本，排除封面、版权、目录、词汇表页和页眉页脚。
    vocab_pages: 词汇表页码范围 (start, end) 1-based，None时自动扫描。
    skip_front_pages: 跳过前N页（封面、版权、目录等），默认8页。
    返回: 句子列表
    """
    sentences = []
    
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        
        # 确定要排除的页面
        exclude_pages = set()
        # 排除前N页（封面、版权、目录）
        for p in range(min(skip_front_pages, total_pages)):
            exclude_pages.add(p)
        # 排除词汇表页
        if vocab_pages:
            for p in range(vocab_pages[0] - 1, vocab_pages[1]):
                exclude_pages.add(p)
        # 自动扫描：排除含"Words and Expressions"或大量音标的页面
        for i in range(total_pages):
            if i in exclude_pages:
                continue
            text = pdf.pages[i].extract_text() or ""
            if re.search(r'Words\s+and\s+Expressions|Vocabulary\s+Index|词汇表|单词表', text, re.IGNORECASE):
                exclude_pages.add(i)
                continue
            # 检测音标特征：ASCII音标（含 @ : {} 等）或 PUA 音标
            ascii_phon_count = len(re.findall(r'/[a-zA-Z@:(){}]+/', text))
            pua_count = len(re.findall(r'[\ue000-\uf8ff]', text))
            if ascii_phon_count >= 10 or pua_count >= 20:
                exclude_pages.add(i)
        
        for i in range(total_pages):
            if i in exclude_pages:
                continue
            page = pdf.pages[i]
            text = page.extract_text()
            if not text:
                continue
            
            # 过滤行
            lines = text.split('\n')
            filtered_lines = []
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                # 排除页码
                if re.fullmatch(r'\d{1,3}', line):
                    continue
                # 排除页眉（Unit N / 标题等）
                if re.match(r'^Unit\s+\d+$', line, re.IGNORECASE):
                    continue
                if re.match(r'^Words\s+and\s+Expressions', line, re.IGNORECASE):
                    continue
                # 排除"关注微信公众号"等广告
                if re.search(r'微信公众号|关注微信|捷思课堂|获取更多学习资料', line):
                    continue
                # 排除单词列表行（单元介绍页的词汇列表）
                if is_word_list_line(line):
                    continue
                # 排除纯中文行（翻译行）
                if re.fullmatch(r'[\u4e00-\u9fff\s，。、；：！？""\'\'（）《》]+', line):
                    continue
                # 排除含特殊符号的行（项目符号等）
                if re.search(r'[\u2460-\u24FF\u00AE\u2122]', line):
                    continue
                filtered_lines.append(line)
            
            if not filtered_lines:
                continue
            
            page_text = ' '.join(filtered_lines)
            
            # 句子分割（.?! 分割，保留标点）
            page_text = re.sub(r'(Mr|Mrs|Ms|Dr|Prof|Sr|Jr|St|vs|etc|e\.g|i\.e|U\.S|U\.K|a\.m|p\.m)\.', r'\1<DOT>', page_text)
            raw_sentences = re.split(r'(?<=[.!?])\s+', page_text)
            
            for s in raw_sentences:
                s = s.replace('<DOT>', '.').strip()
                if is_valid_sentence(s):
                    sentences.append(s)
    
    return sentences


# ==================== 例句匹配 ====================
def find_example_for_word(word, sentences, max_scan=0):
    """为单个单词找高质量例句，返回后处理后的句子或None（None表示需要LLM替换）。
    支持词形变化匹配和PDF拆分单词匹配（如 diff erent → different）。
    匹配后自动执行后处理（修复拆分单词、首字母大写）和质量评估，
    只返回通过质量检查的例句，质量不高的继续找下一个。
    """
    word_lower = word.lower().strip()
    
    # 判断是否是短语（含空格或括号占位符）
    is_phrase = ' ' in word_lower or '(' in word_lower or '...' in word_lower or '⋯' in word_lower
    
    if is_phrase:
        keywords = extract_phrase_keywords(word_lower)
        if not keywords:
            return None
        # 短语匹配：句子中同时出现所有核心关键词（或其变形/拆分形式）
        keyword_patterns = {}
        for kw in keywords:
            forms = generate_word_forms(kw)
            # 为每个形式生成拆分形式，合并成正则模式
            all_forms = set()
            for f in forms:
                all_forms.add(f)
                all_forms.update(generate_split_forms(f))
            # 转义并构建正则（拆分形式中的空格需要匹配任意空白）
            patterns = []
            for f in all_forms:
                escaped = re.escape(f).replace(r'\ ', r'\s+')
                patterns.append(escaped)
            keyword_patterns[kw] = re.compile(r'(?<![a-z])(' + '|'.join(patterns) + r')(?![a-z])', re.IGNORECASE)
        
        scanned = 0
        for s in sentences:
            if max_scan > 0 and scanned >= max_scan:
                break
            all_found = True
            for kw, pat in keyword_patterns.items():
                if not pat.search(s):
                    all_found = False
                    break
            if all_found:
                # 后处理：修复拆分单词、首字母大写
                processed = post_process_example(s)
                # 质量评估：只返回高质量例句
                if is_high_quality_example(processed):
                    return processed
            scanned += 1
        return None
    else:
        # 单词匹配：原形、变形形式、PDF拆分形式
        forms = generate_word_forms(word_lower)
        all_forms = set()
        for f in forms:
            all_forms.add(f)
            all_forms.update(generate_split_forms(f))
        # 构建正则模式（拆分形式中的空格匹配任意空白）
        patterns = []
        for f in all_forms:
            escaped = re.escape(f).replace(r'\ ', r'\s+')
            patterns.append(escaped)
        combined = re.compile(r'(?<![a-z])(' + '|'.join(patterns) + r')(?![a-z])', re.IGNORECASE)
        
        scanned = 0
        for s in sentences:
            if max_scan > 0 and scanned >= max_scan:
                break
            if combined.search(s):
                # 后处理：修复拆分单词、首字母大写
                processed = post_process_example(s)
                # 质量评估：只返回高质量例句
                if is_high_quality_example(processed):
                    return processed
            scanned += 1
        return None


# ==================== LLM 生成兜底 ====================
def generate_examples_with_llm(unmatched_words, api_key, base_url=None, model=None):
    """用 LLM 为未匹配单词批量生成例句。返回 {word: example} 字典。"""
    if not api_key:
        return {}
    
    try:
        import urllib.request
        import json as json_mod
    except ImportError:
        return {}
    
    base_url = base_url or "https://api.openai.com/v1"
    model = model or "gpt-4o-mini"
    
    result = {}
    # 分批处理，每批20个
    batch_size = 20
    for i in range(0, len(unmatched_words), batch_size):
        batch = unmatched_words[i:i+batch_size]
        prompt = f"请为以下初中英语单词/短语各生成一个简单的英文例句（适合初中水平，8-20个单词，必须包含目标单词），以JSON格式返回，key是单词，value是例句。不要输出其他内容。\n\n单词列表：{json.dumps(batch, ensure_ascii=False)}"
        
        try:
            req_body = json_mod.dumps({
                "model": model,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.7,
            }).encode('utf-8')
            
            req = urllib.request.Request(
                f"{base_url}/chat/completions",
                data=req_body,
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                }
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                resp_data = json_mod.loads(resp.read().decode('utf-8'))
                content = resp_data['choices'][0]['message']['content']
                # 解析JSON
                content = content.strip()
                if content.startswith('```'):
                    content = re.sub(r'^```\w*\n', '', content)
                    content = re.sub(r'\n```$', '', content)
                batch_result = json_mod.loads(content)
                result.update(batch_result)
        except Exception as e:
            print(f"LLM 生成失败（批次 {i//batch_size+1}）: {e}", file=sys.stderr)
    
    return result


# ==================== 主流程 ====================
def load_vocab_excel(path):
    """加载词汇表Excel，返回条目列表（字典）。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        return []
    
    headers = [str(h or '').strip().lower() for h in rows[0]]
    # 语义匹配列
    col_map = {}
    for i, h in enumerate(headers):
        if h in ('单词', '英文单词', '英文单词/短语', '英文', 'english', 'word'):
            col_map['word'] = i
        elif h in ('所属单元', '单元', 'unit'):
            col_map['unit'] = i
        elif h in ('词性', 'pos', 'part of speech'):
            col_map['pos'] = i
        elif h in ('中文释义', '释义', '中文', 'meaning', 'definition'):
            col_map['meaning'] = i
        elif h in ('例句', 'example', 'sentence'):
            col_map['example'] = i
    
    entries = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        entry = {}
        for field, idx in col_map.items():
            entry[field] = row[idx] if idx < len(row) else ''
        if 'word' in entry and entry['word']:
            entries.append(entry)
    return entries


def write_output_excel(entries, output_path):
    """写入带例句的Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "词汇表"
    
    headers = ["单词", "所属单元", "词性", "中文释义", "例句"]
    ws.append(headers)
    
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    for e in entries:
        ws.append([
            str(e.get('word', '') or ''),
            str(e.get('unit', '') or ''),
            str(e.get('pos', '') or ''),
            str(e.get('meaning', '') or ''),
            str(e.get('example', '') or ''),
        ])
    
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if c in (2, 3) else "left",
                wrap_text=(c in (4, 5)),
            )
    
    ws.freeze_panes = "A2"
    widths = [32, 12, 16, 46, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="英语教材词汇表例句生成")
    parser.add_argument("vocab", help="词汇表 Excel 文件路径")
    parser.add_argument("pdf", nargs="?", default=None, help="教材 PDF 文件路径（--llm-only模式下可选）")
    parser.add_argument("--output", default=None, help="输出 Excel 路径，默认 <输入名>_带例句.xlsx")
    parser.add_argument("--llm-only", action="store_true", help="全量使用LLM生成例句（跳过教材正文提取）")
    parser.add_argument("--llm-api-key", default=None, help="LLM API key（未匹配单词例句生成兜底，--llm-only模式下必需）")
    parser.add_argument("--llm-base-url", default=None, help="LLM API 基础地址")
    parser.add_argument("--llm-model", default=None, help="LLM 模型名")
    parser.add_argument("--max-scan", type=int, default=0, help="每个单词最多扫描的候选句子数，0表示不限制（默认）")
    parser.add_argument("--json", default=None, help="输出未匹配单词列表（JSON）")
    parser.add_argument("--vocab-pages", default=None, help="词汇表页码范围（如 170-184），不指定则自动扫描排除")
    parser.add_argument("--quiet", action="store_true", help="不输出处理日志")
    args = parser.parse_args()
    
    if not os.path.isfile(args.vocab):
        print(f"ERROR: 词汇表文件不存在: {args.vocab}", file=sys.stderr)
        sys.exit(1)
    if not args.llm_only and not args.pdf:
        print("ERROR: 请指定教材PDF路径，或使用--llm-only全量LLM生成", file=sys.stderr)
        sys.exit(1)
    if not args.llm_only and not os.path.isfile(args.pdf):
        print(f"ERROR: 教材PDF不存在: {args.pdf}", file=sys.stderr)
        sys.exit(1)
    if args.llm_only and not args.llm_api_key:
        print("ERROR: --llm-only模式需要指定--llm-api-key", file=sys.stderr)
        sys.exit(1)
    
    # 输出路径
    if args.output:
        output_path = args.output
    else:
        base, ext = os.path.splitext(args.vocab)
        output_path = f"{base}_带例句.xlsx"
    
    # 解析词汇表页码范围
    vocab_pages = None
    if args.vocab_pages:
        m = re.match(r'(\d+)-(\d+)', args.vocab_pages)
        if m:
            vocab_pages = (int(m.group(1)), int(m.group(2)))
    
    # 1. 加载词汇表
    if not args.quiet:
        print("正在加载词汇表...", file=sys.stderr)
    entries = load_vocab_excel(args.vocab)
    if not args.quiet:
        print(f"  加载 {len(entries)} 条词汇", file=sys.stderr)
    
    # 2. 提取正文句子（--llm-only模式跳过）
    if args.llm_only:
        if not args.quiet:
            print("全量LLM生成模式：跳过教材正文提取", file=sys.stderr)
        matched = 0
        unmatched = [str(e.get('word', '') or '').strip() for e in entries if e.get('word')]
    else:
        if not args.quiet:
            print("正在提取教材正文句子...", file=sys.stderr)
        sentences = extract_body_text(args.pdf, vocab_pages)
        if not args.quiet:
            print(f"  提取 {len(sentences)} 个句子", file=sys.stderr)
        
        # 3. 为每个单词匹配例句
        if not args.quiet:
            print("正在匹配例句...", file=sys.stderr)
        matched = 0
        unmatched = []
        for e in entries:
            word = str(e.get('word', '') or '').strip()
            if not word:
                e['example'] = ''
                continue
            example = find_example_for_word(word, sentences, args.max_scan)
            if example:
                e['example'] = example
                matched += 1
            else:
                e['example'] = ''
                unmatched.append(word)
        
        if not args.quiet:
            print(f"  正文匹配成功: {matched}/{len(entries)} ({matched*100//max(len(entries),1)}%)", file=sys.stderr)
            print(f"  未匹配: {len(unmatched)} 条", file=sys.stderr)
    
    # 4. LLM 生成兜底
    if unmatched and args.llm_api_key:
        if not args.quiet:
            print(f"正在用 LLM 为 {len(unmatched)} 个单词生成例句...", file=sys.stderr)
        llm_examples = generate_examples_with_llm(
            unmatched, args.llm_api_key, args.llm_base_url, args.llm_model
        )
        llm_matched = 0
        for e in entries:
            word = str(e.get('word', '') or '').strip()
            if word in llm_examples and not e.get('example'):
                e['example'] = llm_examples[word]
                llm_matched += 1
        if not args.quiet:
            print(f"  LLM 生成成功: {llm_matched}/{len(unmatched)}", file=sys.stderr)
        # 更新未匹配列表
        unmatched = [w for w in unmatched if w not in llm_examples]
    
    # 5. 输出未匹配列表
    if args.json:
        with open(args.json, 'w', encoding='utf-8') as f:
            json.dump({
                'total': len(entries),
                'matched': matched,
                'unmatched_count': len(unmatched),
                'unmatched_words': unmatched,
            }, f, ensure_ascii=False, indent=2)
        if not args.quiet:
            print(f"  未匹配列表已保存: {args.json}", file=sys.stderr)
    
    # 6. 写入输出
    write_output_excel(entries, output_path)
    
    if not args.quiet:
        final_matched = sum(1 for e in entries if e.get('example'))
        print(f"\n=== 完成 ===", file=sys.stderr)
        print(f"  总词条: {len(entries)}", file=sys.stderr)
        print(f"  有例句: {final_matched} ({final_matched*100//max(len(entries),1)}%)", file=sys.stderr)
        print(f"  无例句: {len(entries) - final_matched}", file=sys.stderr)
        print(f"  输出: {output_path}", file=sys.stderr)


if __name__ == '__main__':
    main()
