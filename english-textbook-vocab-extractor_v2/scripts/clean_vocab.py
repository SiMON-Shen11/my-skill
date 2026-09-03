# -*- coding: utf-8 -*-
"""
英语教材词汇表数据清洗脚本。
接收 extract_vocab.py 输出的词汇表 Excel（或 JSON），执行结构化清洗与转换，
输出符合新格式（单词 | 所属单元 | 词性 | 中文释义 | 例句）的 Excel 文件。

用法:
  python clean_vocab.py <输入文件> [选项]

选项:
  --output PATH        输出 Excel 路径，默认 <输入文件名>_清洗后.xlsx
  --json               输入为 JSON 格式（默认自动检测扩展名）
  --examples PDF       教材PDF路径，清洗后自动为每个单词生成例句（优先从正文提取）
  --vocab-pages RANGE  词汇表页码范围（如 170-184），用于排除词汇表页
  --llm-api-key KEY    LLM API key（未匹配单词的例句生成兜底）
  --llm-base-url URL   LLM API 基础地址
  --llm-model MODEL    LLM 模型名
  --quiet              不输出处理日志

清洗规则:
  1. 列结构：单词 | 所属单元 | 词性 | 中文释义 | 例句
  2. 所属单元：去除所有非数字字符，仅保留阿拉伯数字
  3. 词性标准化：n.→NOUN, v.→VERB, adj.→ADJ, adv.→ADV，含&的拆分后用,连接，其余→other
  4. 新增「例句」列，若指定 --examples 则自动生成例句（优先从教材正文提取，未匹配可用LLM兜底）
  5. 输出仅含一个 sheet
"""
import sys, os, re, json, argparse
from collections import Counter

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def generate_examples_for_entries(entries, pdf_path, vocab_pages=None,
                                   llm_api_key=None, llm_base_url=None, llm_model=None, quiet=False):
    """为清洗后的词条生成例句，直接修改 entries 中的 '例句' 字段。
    返回 (matched_count, total_count, unmatched_words)。
    """
    # 动态导入 generate_examples 模块（同目录）
    import importlib.util
    script_dir = os.path.dirname(os.path.abspath(__file__))
    gen_path = os.path.join(script_dir, 'generate_examples.py')
    if not os.path.isfile(gen_path):
        print("WARNING: generate_examples.py not found, skipping example generation", file=sys.stderr)
        return 0, len(entries), [e['单词'] for e in entries]
    
    spec = importlib.util.spec_from_file_location("generate_examples", gen_path)
    gen_mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(gen_mod)
    
    # 提取正文句子
    if not quiet:
        print("正在提取教材正文句子...", file=sys.stderr)
    sentences = gen_mod.extract_body_text(pdf_path, vocab_pages)
    if not quiet:
        print(f"  提取 {len(sentences)} 个句子", file=sys.stderr)
    
    # 为每个单词匹配例句
    if not quiet:
        print("正在匹配例句...", file=sys.stderr)
    matched = 0
    unmatched = []
    for e in entries:
        word = e['单词']
        example = gen_mod.find_example_for_word(word, sentences, max_scan=0)
        if example:
            e['例句'] = example
            matched += 1
        else:
            unmatched.append(word)
    
    if not quiet:
        print(f"  正文匹配成功: {matched}/{len(entries)} ({matched*100//max(len(entries),1)}%)", file=sys.stderr)
    
    # LLM 生成兜底
    if unmatched and llm_api_key:
        if not quiet:
            print(f"正在用 LLM 为 {len(unmatched)} 个单词生成例句...", file=sys.stderr)
        llm_examples = gen_mod.generate_examples_with_llm(
            unmatched, llm_api_key, llm_base_url, llm_model
        )
        llm_matched = 0
        for e in entries:
            word = e['单词']
            if word in llm_examples and not e['例句']:
                e['例句'] = llm_examples[word]
                llm_matched += 1
        if not quiet:
            print(f"  LLM 生成成功: {llm_matched}/{len(unmatched)}", file=sys.stderr)
        unmatched = [w for w in unmatched if w not in llm_examples]
        matched += llm_matched
    
    return matched, len(entries), unmatched

# ==================== 常量 ====================
POS_MAP = {
    'n.': 'NOUN',
    'v.': 'VERB',
    'adj.': 'ADJ',
    'adv.': 'ADV',
}

# 合法词性基础值
VALID_POS_BASE = {'NOUN', 'VERB', 'ADJ', 'ADV', 'other'}

def is_valid_pos(pos):
    """动态验证词性值：支持任意数量的逗号组合（如 ADJ,ADV,NOUN）。"""
    if not pos:
        return False
    parts = [p.strip() for p in pos.split(',')]
    return all(p in VALID_POS_BASE for p in parts) and len(parts) > 0

OUTPUT_HEADERS = ["单词", "所属单元", "词性", "中文释义", "例句"]

# 列名语义匹配（支持多种命名）
COLUMN_ALIASES = {
    'word': ['单词', '英文单词', '英文单词/短语', '英文', 'english', 'word'],
    'unit': ['所属单元', '单元', 'unit', 'unit n'],
    'pos': ['词性', 'pos', 'part of speech'],
    'meaning': ['中文释义', '释义', '中文', 'meaning', 'definition'],
}


# ==================== 工具函数 ====================
def normalize_pos(pos):
    """标准化词性：含&的拆分映射，其他统一other。"""
    if pos is None:
        return 'other'
    pos = str(pos).strip()
    if not pos:
        return 'other'
    # 含 & 的组合词性
    if '&' in pos:
        parts = [p.strip() for p in pos.split('&')]
        mapped = []
        for p in parts:
            mapped.append(POS_MAP.get(p, 'other'))
        # 去重保持原顺序
        seen = set()
        result = []
        for m in mapped:
            if m not in seen:
                seen.add(m)
                result.append(m)
        return ','.join(result)
    # 单项
    return POS_MAP.get(pos, 'other')


def normalize_unit(unit):
    """去除所有非数字字符，仅保留阿拉伯数字。"""
    if unit is None:
        return '0'
    digits = re.sub(r'[^0-9]', '', str(unit))
    return digits if digits else '0'


def detect_columns(headers):
    """根据表头语义匹配列索引，返回 {field_name: col_index} 或 None。"""
    header_lower = [str(h).strip().lower() if h is not None else '' for h in headers]
    col_map = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_lower = alias.lower()
            for idx, h in enumerate(header_lower):
                if h == alias_lower or alias_lower in h:
                    col_map[field] = idx
                    break
            if field in col_map:
                break
    return col_map if len(col_map) >= 3 else None  # 至少匹配到单词/单元/词性


def load_from_excel(path):
    """从 Excel 文件加载词汇数据，返回条目列表。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # 查找表头行（第一行非空且含关键词）
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        if row and any(h is not None and str(h).strip() for h in row):
            header_idx = i
            break

    headers = rows[header_idx]
    col_map = detect_columns(headers)

    entries = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue  # 跳过空行
        if col_map:
            entry = {}
            for field, idx in col_map.items():
                entry[field] = row[idx] if idx < len(row) else None
            entries.append(entry)
        else:
            # 无法匹配列名时，按位置假设：单词, 单元, 词性, 释义
            if len(row) >= 4:
                entries.append({
                    'word': row[0], 'unit': row[1],
                    'pos': row[2], 'meaning': row[3],
                })
    return entries


def load_from_json(path):
    """从 JSON 文件加载词汇数据（extract_vocab.py 的 --json 输出格式）。"""
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    entries = []
    for item in data:
        entries.append({
            'word': item.get('english', ''),
            'unit': item.get('unit', ''),
            'pos': item.get('pos', ''),
            'meaning': item.get('meaning', ''),
        })
    return entries


def clean_entries(entries):
    """执行清洗，返回 (cleaned_list, stats_dict, other_details)。"""
    cleaned = []
    pos_counter = Counter()
    other_details = []

    for e in entries:
        word = str(e.get('word', '') or '').strip()
        unit_raw = e.get('unit', '')
        pos_raw = e.get('pos', '')
        meaning = str(e.get('meaning', '') or '').strip()

        if not word:
            continue  # 跳过无单词的行

        unit = normalize_unit(unit_raw)
        pos = normalize_pos(pos_raw)

        pos_counter[pos] += 1
        if pos == 'other':
            other_details.append((word, str(pos_raw or '').strip(), meaning[:40]))

        cleaned.append({
            '单词': word,
            '所属单元': unit,
            '词性': pos,
            '中文释义': meaning,
            '例句': '',
        })

    stats = {
        'total': len(cleaned),
        'pos_distribution': dict(pos_counter),
        'other_count': len(other_details),
        'other_details': other_details,
    }
    return cleaned, stats


def self_check(cleaned, source_count):
    """自检，返回 (passed: bool, issues: list)。"""
    issues = []
    # 1. 行数
    if len(cleaned) != source_count:
        issues.append(f"行数不一致: 清洗后{len(cleaned)} vs 源数据{source_count}")
    # 2. 单元纯数字
    bad_units = [e for e in cleaned if not str(e['所属单元']).isdigit()]
    if bad_units:
        issues.append(f"所属单元非纯数字: {len(bad_units)}条，如 {bad_units[0]}")
    # 3. 词性合法（支持任意数量逗号组合）
    bad_pos = [e for e in cleaned if not is_valid_pos(e['词性'])]
    if bad_pos:
        issues.append(f"词性非法: {len(bad_pos)}条，如 {bad_pos[0]}")
    # 4. 无残留 &
    bad_amp = [e for e in cleaned if '&' in str(e['词性'])]
    if bad_amp:
        issues.append(f"残留&: {len(bad_amp)}条")
    return (len(issues) == 0, issues)


def write_excel(cleaned, output_path):
    """写入清洗后的 Excel，仅含一个 sheet。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "词汇表"

    ws.append(OUTPUT_HEADERS)

    # 样式
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, len(OUTPUT_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for e in cleaned:
        ws.append([e['单词'], e['所属单元'], e['词性'], e['中文释义'], e['例句']])

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(OUTPUT_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if c in (2, 3) else "left",
                wrap_text=(c in (4, 5)),
            )

    ws.freeze_panes = "A2"
    widths = [32, 12, 16, 46, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 确保只有一个 sheet
    while len(wb.sheetnames) > 1:
        del wb[wb.sheetnames[-1]]

    wb.save(output_path)
    return len(cleaned)


def print_log(stats, source_count, output_path, check_passed, issues):
    """输出处理日志。"""
    print(f"源数据行数: {source_count}", file=sys.stderr)
    print(f"清洗后行数: {stats['total']}", file=sys.stderr)
    print(f"\n词性分布:", file=sys.stderr)
    for pos, cnt in sorted(stats['pos_distribution'].items(), key=lambda x: -x[1]):
        print(f"  {pos:20s}: {cnt:4d}", file=sys.stderr)
    print(f"\n归为 other 的条目数: {stats['other_count']}", file=sys.stderr)
    if stats['other_details']:
        print(f"other 明细（前10条）:", file=sys.stderr)
        for word, pos_raw, meaning in stats['other_details'][:10]:
            print(f"  单词={word:25s} 原词性={pos_raw:10s} 释义={meaning}", file=sys.stderr)
    print(f"\n自检: {'✓ 通过' if check_passed else '✗ 未通过'}", file=sys.stderr)
    if issues:
        for iss in issues:
            print(f"  - {iss}", file=sys.stderr)
    print(f"\n输出: {output_path} ({stats['total']} rows)", file=sys.stderr)


# ==================== CLI 入口 ====================
def main():
    parser = argparse.ArgumentParser(description="英语教材词汇表数据清洗")
    parser.add_argument("input", help="输入文件路径（Excel .xlsx 或 JSON .json）")
    parser.add_argument("--output", default=None, help="输出 Excel 路径，默认 <输入名>_清洗后.xlsx")
    parser.add_argument("--json", action="store_true", help="强制按 JSON 格式读取输入")
    parser.add_argument("--examples", default=None, help="教材PDF路径，清洗后自动生成例句")
    parser.add_argument("--vocab-pages", default=None, help="词汇表页码范围（如 170-184）")
    parser.add_argument("--llm-api-key", default=None, help="LLM API key（未匹配单词例句生成兜底）")
    parser.add_argument("--llm-base-url", default=None, help="LLM API 基础地址")
    parser.add_argument("--llm-model", default=None, help="LLM 模型名")
    parser.add_argument("--quiet", action="store_true", help="不输出处理日志")
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"ERROR: 输入文件不存在: {args.input}", file=sys.stderr)
        sys.exit(1)

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        base, ext = os.path.splitext(args.input)
        output_path = f"{base}_清洗后.xlsx"

    # 加载数据
    is_json = args.json or args.input.lower().endswith('.json')
    if is_json:
        entries = load_from_json(args.input)
    else:
        entries = load_from_excel(args.input)

    source_count = len(entries)
    if source_count == 0:
        print("ERROR: 未从输入文件中读取到任何词汇数据", file=sys.stderr)
        sys.exit(1)

    # 执行清洗
    cleaned, stats = clean_entries(entries)

    # 生成例句（如果指定了 --examples）
    example_stats = None
    if args.examples:
        if not os.path.isfile(args.examples):
            print(f"ERROR: 教材PDF不存在: {args.examples}", file=sys.stderr)
            sys.exit(1)
        # 解析词汇表页码范围
        vocab_pages = None
        if args.vocab_pages:
            m = re.match(r'(\d+)-(\d+)', args.vocab_pages)
            if m:
                vocab_pages = (int(m.group(1)), int(m.group(2)))
        example_stats = generate_examples_for_entries(
            cleaned, args.examples, vocab_pages,
            args.llm_api_key, args.llm_base_url, args.llm_model, args.quiet
        )
        if not args.quiet:
            matched, total, unmatched = example_stats
            print(f"  例句覆盖率: {matched}/{total} ({matched*100//max(total,1)}%)", file=sys.stderr)
            if unmatched:
                print(f"  未匹配单词: {len(unmatched)} 个（可手动补充或配置LLM API key）", file=sys.stderr)

    # 自检
    check_passed, issues = self_check(cleaned, source_count)

    # 写入 Excel
    write_excel(cleaned, output_path)

    # 输出日志
    if not args.quiet:
        print_log(stats, source_count, output_path, check_passed, issues)

    # 自检未通过时返回非零退出码（但不中断，文件已生成）
    if not check_passed:
        sys.exit(2)


if __name__ == '__main__':
    main()
